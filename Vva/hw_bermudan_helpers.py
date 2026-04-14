"""
hw_bermudan_helpers.py
======================
Helper module for the Bermudan Swaption CVA notebook.

Implements She & Grecu (2018), arXiv:1811.08726, Section 3.

Contains every class and function used in the notebook — nothing executes
on import. The notebook calls these in order, one cell per stage.

Contents
--------
  VolSurface              — piecewise-constant σ_r(t) term structure
  Config                  — all parameters in one dataclass
  HullWhite               — ZCB prices, swap values, numeraire (uses VolSurface)
  ValueNetwork            — MLP: (t, x) → Ṽ(t, x)
  pre_training()          — Stage 1: simulate x(t) paths + exercise values
  _apply_exercise_jumps() — jump condition Ṽ(T_m^-) = max(Ṽ(T_m^+), Ũ(T_m))
  _build_targets()        — Picard training targets from jump-corrected Y
  _optimise()             — Adam training loop
  _evaluate_network()     — evaluate Y_net on all paths × all time steps
  train()                 — Stage 2: Picard iteration with exercise jumps
  compute_exposure()      — Stage 3: EPE, ENE, CVA, exercise time artifacts
  bachelier_formula()     — eq (3.8) option price function
  fit_bachelier()         — scipy curve_fit wrapper
  plot_epe_ene()          — Fig 3.1 left
  plot_loss_curve()       — Fig 3.1 right
  plot_future_value_evolution() — Fig 3.2 / 3.3
  plot_bachelier_fit()    — Fig 3.4
  plot_exercise_time_distribution() — exercise artifacts
  export_to_excel()       — 5-sheet formatted workbook
"""

# ─────────────────────────────────────────────────────────────────────────────
# Imports
# ─────────────────────────────────────────────────────────────────────────────

import math
import logging
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Union

import numpy as np
import torch
import torch.nn as nn
from scipy.optimize import curve_fit
from scipy.stats import norm as scipy_norm
import matplotlib.pyplot as plt
import pandas as pd

# Module-level logger — the notebook configures the root logger level
logger = logging.getLogger(__name__)

# Device is resolved once at import time so every function uses the same one
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")




# ─────────────────────────────────────────────────────────────────────────────
# 0b. Vol Surface — Piecewise-Constant σ_r(t)
# ─────────────────────────────────────────────────────────────────────────────

class VolSurface:
    """
    Piecewise-constant Hull-White short-rate volatility term structure σ_r(t).

    In a calibrated model (She & Grecu Section 3.2: "calibrated to market data
    on 1/18/2018"), σ_r is not a single flat constant but a step function
    bootstrapped from market swaption implied volatilities.

    Representation
    --------------
    The surface is defined by a list of (t_i, σ_i) breakpoints where σ_i
    applies on the half-open interval [t_{i-1}, t_i):

        σ_r(t) = σ_i    for  t ∈ [t_{i-1}, t_i),   i = 1, …, n
        σ_r(t) = σ_n    for  t ≥ t_n   (flat extrapolation beyond last knot)

    The first breakpoint t_0 = 0 is implicit — you only supply the right
    endpoints of each piece.

    Example
    -------
    Breakpoints covering 0–7.5y with three calibration instruments:

        vol = VolSurface(
            breakpoints = [1.5, 3.5, 7.5],
            sigmas      = [0.008, 0.011, 0.013],
        )

    Flat default (equivalent to the old scalar sigma_r):

        vol = VolSurface.flat(0.01)

    Key methods
    -----------
    sigma_at(t)                  : σ_r(t) — piecewise lookup
    integrated_variance(t)       : ∫₀ᵗ σ_r(u)² du — for ln_A convexity
    weighted_integrated_variance : ∫₀ᵗ e^{-2κ(t-u)} σ_r(u)² du — for y(t)
    """

    def __init__(self, breakpoints: List[float], sigmas: List[float]):
        """
        Args:
            breakpoints : right endpoints of each piecewise-constant segment,
                          strictly increasing. E.g. [1.0, 2.0, 5.0, 10.0].
            sigmas      : σ_r value on each segment. len(sigmas) == len(breakpoints).
                          Extrapolated flat beyond the final breakpoint.
        """
        if len(breakpoints) != len(sigmas):
            raise ValueError(
                f"breakpoints and sigmas must have equal length, "
                f"got {len(breakpoints)} and {len(sigmas)}."
            )
        if any(b <= 0 for b in breakpoints):
            raise ValueError("All breakpoints must be positive.")
        if list(breakpoints) != sorted(breakpoints):
            raise ValueError("breakpoints must be strictly increasing.")
        if any(s <= 0 for s in sigmas):
            raise ValueError("All sigma values must be positive.")

        # Left edges: [0, t1, t2, …, tn]
        self._bp  = np.array([0.0] + list(breakpoints))
        self._sig = np.array(sigmas, dtype=float)

        # Precompute cumulative ∫ σ² du at each breakpoint for fast lookup
        n = len(self._sig)
        self._cum_iv = np.zeros(n + 1)
        for k in range(n):
            self._cum_iv[k + 1] = (
                self._cum_iv[k] + self._sig[k]**2 * (self._bp[k + 1] - self._bp[k])
            )

    @classmethod
    def flat(cls, sigma: float, horizon: float = 100.0) -> "VolSurface":
        """
        Convenience constructor for a flat (constant) volatility.
        Equivalent to the legacy scalar sigma_r parameter.

            vol = VolSurface.flat(0.01)   # same as sigma_r = 0.01
        """
        return cls(breakpoints=[horizon], sigmas=[sigma])

    def sigma_at(self, t: float) -> float:
        """
        σ_r(t) — piecewise-constant lookup.

        Returns σ_i for t in [t_{i-1}, t_i). Extrapolates flat beyond
        the last breakpoint.

        Args:
            t : time (scalar, t ≥ 0)
        """
        idx = int(np.clip(
            np.searchsorted(self._bp, t, side="right") - 1,
            0, len(self._sig) - 1,
        ))
        return float(self._sig[idx])

    def integrated_variance(self, t: float) -> float:
        """
        ∫₀ᵗ σ_r(u)² du  — analytically, summing over piecewise segments.

        Used in the convexity correction term of ln_A(t, T):
            -0.5 * B(t,T)² * ∫₀ᵗ σ_r(u)² du   (Andersen & Piterbarg, Vol 2)

        Note: for a flat surface this reduces to σ_r² * t.
        """
        if t <= 0.0:
            return 0.0
        idx = int(np.clip(
            np.searchsorted(self._bp, t, side="right") - 1,
            0, len(self._sig) - 1,
        ))
        iv = float(self._cum_iv[idx])
        iv += self._sig[idx]**2 * (t - self._bp[idx])
        return iv

    def weighted_integrated_variance(self, kappa: float, t: float) -> float:
        """
        y(t) = ∫₀ᵗ e^{-2κ(t-u)} σ_r(u)² du  — analytically.

        This is the HW variance / drift correction function. For each
        piecewise segment [t_{i-1}, t_i) with constant σ_i:

            contribution = σ_i² / (2κ) * [e^{-2κ(t-t_i)} - e^{-2κ(t-t_{i-1})}]

        In the κ → 0 limit: contribution = σ_i² * (t_i - t_{i-1}).

        Args:
            kappa : mean-reversion speed κ
            t     : upper limit (t ≥ 0)
        """
        if t <= 0.0:
            return 0.0

        result = 0.0
        n      = len(self._sig)

        for i in range(n):
            t_lo = self._bp[i]
            t_hi = min(self._bp[i + 1] if i < n - 1 else t, t)

            if t_lo >= t:
                break
            if t_hi <= t_lo:
                continue

            sig2 = self._sig[i]**2
            if kappa < 1e-12:
                result += sig2 * (t_hi - t_lo)
            else:
                result += (sig2 / (2 * kappa)) * (
                    math.exp(-2 * kappa * (t - t_hi))
                    - math.exp(-2 * kappa * (t - t_lo))
                )
        return result

    def __repr__(self) -> str:
        pieces = [
            f"[{self._bp[i]:.2f}, {self._bp[i+1]:.2f}): σ={self._sig[i]:.5f}"
            for i in range(len(self._sig))
        ]
        return "VolSurface(\n  " + "\n  ".join(pieces) + "\n)"

# ─────────────────────────────────────────────────────────────────────────────
# 1.  Configuration
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class Config:
    """
    All model, product, and training parameters in one dataclass.

    Defaults match Section 3.2 of She & Grecu (2018):
        κ=0.01, σ_r=0.01, r0=2.8%, notional=10,000, K=2.8%,
        swap tenor 4y, semi-annual exercise from 1.5y to 3.5y.

    Change any field here before passing to pre_training() / train().
    Derived fields (T, N, exercise_dates) are computed automatically.
    """

    # ── Hull-White 1-factor model ─────────────────────────────────────────────
    kappa:   float = 0.01    # mean-reversion speed κ

    # sigma_r accepts either:
    #   float      — flat constant (legacy default, e.g. 0.01)
    #   VolSurface — piecewise-constant calibrated term structure
    # In both cases HullWhite normalises to a VolSurface internally.
    sigma_r: Union[float, "VolSurface"] = 0.01

    r0:      float = 0.028   # flat initial forward rate f(0,t) = r0

    # ── Underlying swap ───────────────────────────────────────────────────────
    notional:   float = 10_000.0  # notional N
    fixed_rate: float = 0.028     # fixed coupon K (set = r0 for ATM)
    swap_tenor: float = 4.0       # swap length in years from each exercise date
    fixed_freq: float = 0.5       # fixed leg payment frequency (semi-annual)
    float_freq: float = 0.25      # floating leg reset frequency (3M Libor)

    # ── Exercise schedule ─────────────────────────────────────────────────────
    exercise_start: float = 1.5   # first exercise date (years)
    exercise_end:   float = 3.5   # last exercise date
    exercise_freq:  float = 0.5   # semi-annual step

    # ── Monte Carlo simulation ────────────────────────────────────────────────
    num_paths: int   = 5_000      # number of Monte Carlo paths M
    dt:        float = 1 / 52     # ~weekly time steps (~52 per year)

    # ── Neural network architecture ───────────────────────────────────────────
    # Paper Section 3.2: 2 hidden layers, d̃=10 → width = d + d̃ = 1+10 = 11
    hidden_dim:        int = 11   # neurons per hidden layer
    num_hidden_layers: int = 2    # number of hidden layers

    # ── Training ──────────────────────────────────────────────────────────────
    num_train_steps: int   = 5_000   # Adam gradient steps per Picard iteration
    learning_rate:   float = 1e-3    # initial Adam learning rate
    lr_decay_steps:  int   = 1_000   # StepLR period (steps between LR decays)
    lr_decay_rate:   float = 0.5     # StepLR multiplicative factor
    log_every:       int   = 500     # print training loss every N steps

    # ── CVA credit parameters ─────────────────────────────────────────────────
    hazard_rate: float = 0.01   # constant hazard rate λ
    recovery:    float = 0.40   # recovery rate R

    # ── Derived fields (populated automatically in __post_init__) ─────────────
    T:              float       = field(init=False)   # total simulation horizon
    N:              int         = field(init=False)   # number of time steps
    exercise_dates: List[float] = field(init=False)   # [T_1, ..., T_M]

    def __post_init__(self):
        # Simulation horizon = last exercise + full swap tenor that follows it
        self.T = self.exercise_end + self.swap_tenor

        # Adjust N so that dt divides T exactly (avoids index misalignment)
        self.N  = max(1, round(self.T / self.dt))
        self.dt = self.T / self.N

        # Semi-annual exercise dates from exercise_start to exercise_end
        n_ex = round((self.exercise_end - self.exercise_start) / self.exercise_freq) + 1
        self.exercise_dates = [
            round(self.exercise_start + i * self.exercise_freq, 10)
            for i in range(n_ex)
        ]


# ─────────────────────────────────────────────────────────────────────────────
# 2.  Hull-White 1-Factor Model
# ─────────────────────────────────────────────────────────────────────────────

class HullWhite:
    """
    Analytical zero-coupon bond prices and swap values under HW 1-factor.

    Short rate:   r(t) = x(t) + f(0,t),   f(0,t) = r0 (flat curve)
    Stochastic:   dx(t) = [y(t) - κ x(t)] dt + σ_r(t) dW(t),   x(0) = 0

    σ_r(t) is accepted as either a flat float or a calibrated VolSurface.
    In both cases it is normalised internally to a VolSurface, so the rest
    of the class always calls self.vol.sigma_at(t) / self.vol.weighted_integrated_variance().

    All methods that take x operate on numpy arrays of shape [M].

    References
    ----------
    Andersen & Piterbarg (2010), Vol 2, Chapter 10 — full piecewise formulae.
    She & Grecu (2018), eqs (3.5)–(3.6).
    """

    def __init__(self, cfg: Config):
        self.kappa = cfg.kappa
        self.r0    = cfg.r0

        # Normalise sigma_r to a VolSurface regardless of input type.
        # A plain float becomes a single-segment flat surface.
        if isinstance(cfg.sigma_r, (int, float)):
            self.vol = VolSurface.flat(float(cfg.sigma_r))
        else:
            self.vol = cfg.sigma_r   # already a VolSurface

    # ── σ_r(t) access ────────────────────────────────────────────────────────

    def sigma_at(self, t: float) -> float:
        """
        σ_r(t) — delegates to VolSurface.sigma_at().
        Used in the Euler-Maruyama diffusion step.
        """
        return self.vol.sigma_at(t)

    # ── Scalar analytical helpers ─────────────────────────────────────────────

    def y(self, t: float) -> float:
        """
        HW variance / drift function:
            y(t) = ∫₀ᵗ e^{-2κ(t-u)} σ_r(u)² du

        For a flat σ_r this reduces to (σ_r²/2κ)(1 - e^{-2κt}).
        For a piecewise σ_r(u) it is computed analytically by VolSurface,
        summing one closed-form contribution per segment.

        This appears in the HW drift: dx = [y(t) - κ x] dt + σ_r(t) dW.
        """
        return self.vol.weighted_integrated_variance(self.kappa, t)

    def B(self, t: float, T: float) -> float:
        """
        Bond duration factor B(t,T) = (1/κ)(1 - e^{-κ(T-t)}).
        Used in the ZCB formula: P(t,T) = A(t,T) * exp(-B(t,T) * x(t)).
        This is independent of σ_r(t).
        """
        tau = T - t
        if self.kappa < 1e-12:
            return tau
        return (1.0 - math.exp(-self.kappa * tau)) / self.kappa

    def ln_A(self, t: float, T: float) -> float:
        """
        Log of the A(t,T) factor in the zero-coupon bond formula.

        General form (Andersen & Piterbarg, Vol 2, eq 10.36):
            ln A(t,T) = ln[P(0,T)/P(0,t)] + B(t,T)*f(0,t) - 0.5*B(t,T)²*IV(t)

        where IV(t) = ∫₀ᵗ σ_r(u)² du  (plain integrated variance, not weighted).

        For a flat curve f(0,t) = r0:
            ln[P(0,T)/P(0,t)] = -r0*(T-t)
            B(t,T)*f(0,t)     = B(t,T)*r0
            convexity         = -0.5 * B(t,T)² * IV(t)

        Note: The convexity term uses ∫ σ² du (not the weighted y(t)).
        Both are equal for flat σ_r; they differ for a term structure.
        """
        Bval   = self.B(t, T)
        log_zc = -self.r0 * (T - t)                        # flat ZCB ratio
        drift  = Bval * self.r0                             # f(0,t) adjustment
        IV_t   = self.vol.integrated_variance(t)            # ∫₀ᵗ σ_r(u)² du
        convex = -0.5 * Bval**2 * IV_t                     # convexity correction
        return log_zc + drift + convex

    # ── Vectorised bond price ─────────────────────────────────────────────────

    def zcb(self, t: float, T: float, x: np.ndarray) -> np.ndarray:
        """
        Zero-coupon bond price at time t for maturity T, given x(t):
            P(t, T | x) = exp(ln A(t,T) - B(t,T) * x)

        Args:
            t : current time (scalar)
            T : bond maturity (scalar, must satisfy T > t)
            x : short-rate factor x(t), shape [M]

        Returns:
            P : bond prices, shape [M]
        """
        return np.exp(self.ln_A(t, T) - self.B(t, T) * x)

    # ── Swap and swaption exercise value ──────────────────────────────────────

    def receiver_swap_value(self, t_exercise: float, x: np.ndarray,
                            cfg: Config) -> np.ndarray:
        """
        Analytical value of a *receiver* swap starting at t_exercise.

        Receiver swap: receive fixed coupon K, pay floating (Libor).
            V_receiver = fixed_leg - float_leg
            fixed_leg  = N * K * Σ_k τ * P(t, T_k)
            float_leg  = N * [P(t, T_start) - P(t, T_end)]

        The float leg simplification uses the fact that a floating bond
        prices at par at each reset date, so V_float = N*(P(t,T_start) - P(t,T_end)).

        Args:
            t_exercise : T_m — exercise date, also the swap start (scalar)
            x          : x(T_m), shape [M]
            cfg        : Config

        Returns:
            V_receiver : shape [M]  (positive = in-the-money for receiver)
        """
        N_not  = cfg.notional
        K      = cfg.fixed_rate
        tau    = cfg.fixed_freq
        n_fix  = round(cfg.swap_tenor / cfg.fixed_freq)

        # Absolute fixed payment dates: T_m + k*fixed_freq, k=1,...,n_fix
        fixed_dates = [t_exercise + (i + 1) * cfg.fixed_freq for i in range(n_fix)]

        # Fixed leg: N * K * Σ τ * P(t, T_k)
        fixed_leg = sum(tau * self.zcb(t_exercise, Tk, x) for Tk in fixed_dates)
        fixed_leg *= N_not * K

        # Float leg: N * [P(t, T_start) - P(t, T_end)]
        # P(t_exercise, t_exercise) = 1 by definition (same date)
        P_end     = self.zcb(t_exercise, fixed_dates[-1], x)
        float_leg = N_not * (1.0 - P_end)

        return fixed_leg - float_leg

    def exercise_value(self, t_exercise: float, x: np.ndarray,
                       cfg: Config) -> np.ndarray:
        """
        Cash-settled exercise value of a *receiver* swaption at T_m:
            U(T_m) = max(V_receiver(T_m, x), 0)

        ITM when fixed rate K > current swap rate (high rates environment).

        Args:
            t_exercise : T_m (scalar)
            x          : x(T_m), shape [M]
            cfg        : Config

        Returns:
            U : shape [M], non-negative
        """
        return np.maximum(self.receiver_swap_value(t_exercise, x, cfg), 0.0)


# ─────────────────────────────────────────────────────────────────────────────
# 3.  Neural Network — Y_net(t, x) → Ṽ(t, x)
# ─────────────────────────────────────────────────────────────────────────────

class ValueNetwork(nn.Module):
    """
    MLP approximating the discounted future value Ṽ(t, x) = V(t, x) / B(t).

    Architecture (paper eq 3.7, Section 3.2):
        Input  : [t, x]           — 2D concatenated input
        Hidden : num_hidden layers of width hidden_dim, tanh activation
        Output : scalar Ṽ         — discounted portfolio value

    Note on parameterisation choice
    --------------------------------
    The paper parameterises the *Delta* ∂Ṽ/∂x with one MLP per time step
    (the Han/Jentzen/E architecture). Here we parameterise Ṽ(t, x) directly
    with a single shared MLP — the Picard approach used in the thesis.
    Both solve the same BSDE; the difference is in what the NN approximates.
    """

    def __init__(self, hidden_dim: int = 11, num_hidden: int = 2):
        super().__init__()

        layers: List[nn.Module] = []
        in_dim = 2   # input: (t, x)

        for _ in range(num_hidden):
            layers.append(nn.Linear(in_dim, hidden_dim))
            layers.append(nn.Tanh())
            in_dim = hidden_dim

        layers.append(nn.Linear(in_dim, 1))   # scalar output
        self.net = nn.Sequential(*layers)

        # Small uniform weight initialisation — mirrors random start in paper
        for m in self.modules():
            if isinstance(m, nn.Linear):
                nn.init.uniform_(m.weight, -0.1, 0.1)
                nn.init.zeros_(m.bias)

    def forward(self, t_x: torch.Tensor) -> torch.Tensor:
        """
        Args:
            t_x : [M, 2] — concatenated (t, x) inputs
        Returns:
            V_tilde : [M, 1] — discounted value estimate
        """
        return self.net(t_x)


# ─────────────────────────────────────────────────────────────────────────────
# 4.  Stage 1 — Pre-Training
# ─────────────────────────────────────────────────────────────────────────────

def pre_training(cfg: Config, hw: HullWhite) -> Dict:
    """
    Stage 1 of She & Grecu (2018) Section 3.1.

    (a) Simulate x(t) paths via Euler-Maruyama and store dW increments.
    (b) At each exercise date T_m, compute the undiscounted exercise value
        U(T_m) = max(V_receiver_swap, 0) analytically from ZCB prices.
    (c) Compute the numeraire B(T_m) = exp(∫₀^{T_m} r(s) ds) along each
        path via the trapezoidal rule, then form Ũ(T_m) = U(T_m) / B(T_m).

    Args:
        cfg : Config
        hw  : HullWhite instance

    Returns
    -------
    dict with keys:
        'time_grid'   : [N+1]       uniform time grid 0 → T
        'X_paths'     : [M, N+1]    stochastic factor x(t) per path
        'dW'          : [M, N]      Brownian increments √dt * Z
        'U_exercise'  : [M, num_ex] undiscounted exercise values U(T_m)
        'U_tilde'     : [M, num_ex] discounted exercise values Ũ = U/B
        'B_exercise'  : [M, num_ex] numeraire B(T_m) per path
        'ex_step_idx' : [num_ex]    time-step indices for exercise dates
    """
    logger.info("=" * 60)
    logger.info("STAGE 1 — Pre-training: simulating paths & exercise values")

    M  = cfg.num_paths
    N  = cfg.N
    dt = cfg.dt

    # ── (a)  Euler-Maruyama: dx = [y(t) - κ x] dt + σ_r dW ──────────────────
    time_grid = np.linspace(0.0, cfg.T, N + 1)        # shape [N+1]
    X_paths   = np.zeros((M, N + 1))                  # x(0) = 0
    dW        = np.random.randn(M, N) * math.sqrt(dt) # Brownian increments [M,N]

    for i in range(N):
        t_i    = time_grid[i]
        drift  = (hw.y(t_i) - cfg.kappa * X_paths[:, i]) * dt
        # Use hw.sigma_at(t_i) so that a calibrated VolSurface is respected.
        # For a flat surface this is identical to the old cfg.sigma_r constant.
        sig_ti = hw.sigma_at(t_i)
        X_paths[:, i + 1] = X_paths[:, i] + drift + sig_ti * dW[:, i]

    logger.info(f"  Simulated {M:,} paths, {N} steps (T={cfg.T:.1f}y, dt={dt:.4f})")

    # ── (b)  Map exercise dates T_m to their nearest time-step index ──────────
    ex_step_idx = [min(int(round(T_m / dt)), N) for T_m in cfg.exercise_dates]

    # ── (c)  Exercise values U(T_m) and numeraire B(T_m) ─────────────────────
    num_ex     = len(cfg.exercise_dates)
    U_exercise = np.zeros((M, num_ex))
    U_tilde    = np.zeros((M, num_ex))
    B_exercise = np.zeros((M, num_ex))

    for m_idx, (T_m, step_m) in enumerate(zip(cfg.exercise_dates, ex_step_idx)):
        x_Tm = X_paths[:, step_m]

        # Undiscounted exercise value (analytical, no NN involved)
        U_exercise[:, m_idx] = hw.exercise_value(T_m, x_Tm, cfg)

        # Numeraire B(T_m) = exp(∫₀^{T_m} r(s) ds), r(s) = x(s) + r0
        r_path_m = X_paths[:, :step_m + 1] + cfg.r0    # [M, step_m+1]
        integral = np.trapezoid(r_path_m, dx=dt, axis=1) # [M]
        B_m      = np.exp(integral)
        B_exercise[:, m_idx] = B_m

        # Discounted exercise value: Ũ(T_m) = U(T_m) / B(T_m)
        U_tilde[:, m_idx] = U_exercise[:, m_idx] / B_m

        logger.info(
            f"  T_m={T_m:.1f}y (step {step_m}): "
            f"mean U={np.mean(U_exercise[:, m_idx]):.2f}, "
            f"mean Ũ={np.mean(U_tilde[:, m_idx]):.4f}, "
            f"% ITM={100 * np.mean(U_exercise[:, m_idx] > 0):.1f}%"
        )

    logger.info("Stage 1 complete.\n")
    return {
        'time_grid':   time_grid,
        'X_paths':     X_paths,
        'dW':          dW,
        'U_exercise':  U_exercise,
        'U_tilde':     U_tilde,
        'B_exercise':  B_exercise,
        'ex_step_idx': ex_step_idx,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 5.  Stage 2 — Training Internals (private helpers)
# ─────────────────────────────────────────────────────────────────────────────

def _apply_exercise_jumps(
    Y_tilde: np.ndarray,
    U_tilde: np.ndarray,
    ex_step_idx: List[int],
) -> np.ndarray:
    """
    Apply the early-exercise jump condition (paper eq 2.14) to a Y path:
        Ṽ(T_m^-) = max(Ṽ(T_m^+),  Ũ(T_m))

    The "^+" value is the network's raw continuation estimate at T_m.
    The "^-" value is the holding value after the option decision is made.
    For a cash-settled option the holder exercises whenever U > continuation.

    This function is called in TWO places:
        - inside train() to build jump-corrected targets for the network
        - inside compute_exposure() to get the correct V path for EPE/ENE
    Consistency in both places is essential — see design discussion.

    Args:
        Y_tilde     : [M, N+1]   raw network output (or previous iterate)
        U_tilde     : [M, num_ex] discounted exercise values Ũ(T_m)
        ex_step_idx : [num_ex]   time-step index of each exercise date

    Returns:
        Y_jumped : [M, N+1]  copy with jumps applied at exercise steps
    """
    Y = Y_tilde.copy()
    for m_idx, step_m in enumerate(ex_step_idx):
        # At step_m: override continuation with max(continuation, exercise)
        Y[:, step_m] = np.maximum(Y[:, step_m], U_tilde[:, m_idx])
    return Y


def _build_targets(
    Y_tilde_jumped: np.ndarray,
    time_grid: np.ndarray,
    X_paths: np.ndarray,
    cfg: Config,
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Build supervised (input, target) pairs for Y_net from jump-corrected Y.

    Since the BSDE driver f = 0 (Ṽ is a Q-martingale), the Feynman-Kac
    target for Ṽ(t_i, x_i) collapses to the next-step value:
        target(t_i, x_i) = Ṽ_jumped(t_{i+1})

    We stack all time steps and all paths into a single flat dataset,
    which is then fed to the Adam optimiser in mini-batches.

    Args:
        Y_tilde_jumped : [M, N+1]  jump-corrected discounted value path
        time_grid      : [N+1]
        X_paths        : [M, N+1]
        cfg            : Config

    Returns:
        inputs  : [M*N, 2]  — (t_i, x_i) pairs
        targets : [M*N, 1]  — Ṽ_jumped(t_{i+1})
    """
    M = cfg.num_paths
    N = cfg.N

    inputs_list, targets_list = [], []
    for i in range(N):
        t_vec = np.full(M, time_grid[i])                     # [M]
        x_i   = X_paths[:, i]                                # [M]
        tgt   = Y_tilde_jumped[:, i + 1]                     # [M]

        inputs_list.append(np.stack([t_vec, x_i], axis=1))  # [M, 2]
        targets_list.append(tgt[:, None])                    # [M, 1]

    return (
        np.concatenate(inputs_list,  axis=0),   # [M*N, 2]
        np.concatenate(targets_list, axis=0),   # [M*N, 1]
    )


def _optimise(
    net: nn.Module,
    inputs_np: np.ndarray,
    targets_np: np.ndarray,
    cfg: Config,
) -> List[float]:
    """
    Adam training loop with StepLR learning-rate decay.

    Runs cfg.num_train_steps gradient steps on random mini-batches of
    size 2048 drawn from the full (t, x) dataset.

    Args:
        net        : ValueNetwork (modified in-place)
        inputs_np  : [M*N, 2]  float32 numpy
        targets_np : [M*N, 1]  float32 numpy

    Returns:
        losses : list of per-step MSE loss values (length = num_train_steps)
    """
    inputs  = torch.tensor(inputs_np,  dtype=torch.float32, device=DEVICE)
    targets = torch.tensor(targets_np, dtype=torch.float32, device=DEVICE)

    optimiser = torch.optim.Adam(net.parameters(), lr=cfg.learning_rate)
    scheduler = torch.optim.lr_scheduler.StepLR(
        optimiser, step_size=cfg.lr_decay_steps, gamma=cfg.lr_decay_rate
    )
    loss_fn      = nn.MSELoss()
    losses       = []
    dataset_size = inputs.shape[0]
    indices      = torch.arange(dataset_size, device=DEVICE)

    net.train()
    for step in range(cfg.num_train_steps):
        # Random mini-batch of 2048 samples from the full dataset
        perm      = indices[torch.randperm(dataset_size, device=DEVICE)]
        batch_in  = inputs[perm[:2048]]
        batch_tgt = targets[perm[:2048]]

        optimiser.zero_grad()
        pred = net(batch_in)
        loss = loss_fn(pred, batch_tgt)
        loss.backward()
        torch.nn.utils.clip_grad_norm_(net.parameters(), 1.0)   # prevent exploding grads
        optimiser.step()
        scheduler.step()
        losses.append(loss.item())

        if (step + 1) % cfg.log_every == 0:
            logger.info(f"    step {step+1:5d}/{cfg.num_train_steps}:  loss={loss.item():.4e}")

    return losses


def _evaluate_network(
    net: nn.Module,
    time_grid: np.ndarray,
    X_paths: np.ndarray,
    cfg: Config,
) -> np.ndarray:
    """
    Evaluate Y_net(t_i, x_i) for every path and every time step.

    Pins the terminal value to Ṽ(T_N) = 0 (paper eq 2.16: the option
    value is zero right after the last possible exercise date).

    Args:
        net       : ValueNetwork (eval mode applied internally)
        time_grid : [N+1]
        X_paths   : [M, N+1]
        cfg       : Config

    Returns:
        Y_tilde : [M, N+1]  raw network output (no exercise jumps yet)
    """
    M = cfg.num_paths
    N = cfg.N

    net.eval()
    Y_tilde = np.zeros((M, N + 1))

    with torch.no_grad():
        for i in range(N):
            t_vec = np.full(M, time_grid[i])
            nn_in = torch.tensor(
                np.stack([t_vec, X_paths[:, i]], axis=1),
                dtype=torch.float32, device=DEVICE,
            )
            Y_tilde[:, i] = net(nn_in).squeeze(-1).cpu().numpy()

    # Terminal boundary condition (eq 2.16): Ṽ(T_N^+) = 0
    Y_tilde[:, N] = 0.0
    return Y_tilde


# ─────────────────────────────────────────────────────────────────────────────
# 6.  Stage 2 — Public Training Function
# ─────────────────────────────────────────────────────────────────────────────

def train(
    cfg: Config,
    hw: HullWhite,
    pre_train_data: Dict,
    num_picard_iters: int = 3,
) -> Tuple[nn.Module, List[float], List[np.ndarray]]:
    """
    Stage 2 of She & Grecu (2018) Section 3.1.

    Picard fixed-point iteration with early-exercise jump:

        For k = 1, ..., num_picard_iters:
            (a) Evaluate Y_net on all paths         → Y_raw [M, N+1]
            (b) Apply exercise jumps                 → Y_jumped [M, N+1]
            (c) Build training targets from Y_jumped → (inputs, targets)
            (d) Retrain Y_net via Adam               → updated weights

    The jump in step (b) enforces the early-exercise condition
        Ṽ(T_m^-) = max(Ṽ(T_m^+), Ũ(T_m))
    so the network learns to reproduce a value function that respects
    optimal stopping.

    Args:
        cfg              : Config
        hw               : HullWhite
        pre_train_data   : output of pre_training()
        num_picard_iters : number of Picard iterations (default 3)

    Returns
    -------
    net         : trained ValueNetwork
    all_losses  : flat list of MSE losses across all Picard iterations
    Y_snapshots : list of jump-corrected Y arrays after each iteration
                  (used for Fig 3.2 / 3.3 — value evolution with training)
    """
    logger.info("=" * 60)
    logger.info("STAGE 2 — Training: Picard iteration with exercise jumps")

    time_grid   = pre_train_data['time_grid']
    X_paths     = pre_train_data['X_paths']
    U_tilde     = pre_train_data['U_tilde']
    ex_step_idx = pre_train_data['ex_step_idx']

    # Fresh network — weights initialised uniformly in [-0.1, 0.1]
    net = ValueNetwork(
        hidden_dim=cfg.hidden_dim,
        num_hidden=cfg.num_hidden_layers,
    ).to(DEVICE)

    all_losses  = []
    Y_snapshots = []   # one entry per Picard iteration + final

    for k in range(num_picard_iters):
        logger.info(f"\n  Picard iteration {k+1}/{num_picard_iters}")

        # (a) Raw network evaluation on all paths
        Y_raw    = _evaluate_network(net, time_grid, X_paths, cfg)

        # (b) Apply early-exercise jumps to get the held value
        Y_jumped = _apply_exercise_jumps(Y_raw, U_tilde, ex_step_idx)
        Y_snapshots.append(Y_jumped.copy())

        # Paper eq (3.1) diagnostic: variance of Ṽ(t=0) across paths
        # At convergence all paths should give the same V0, so variance → 0
        V0_paths   = Y_jumped[:, 0]
        paper_loss = float(np.var(V0_paths))
        logger.info(
            f"  Paper loss var(Ṽ_p(0)) = {paper_loss:.4e} | "
            f"mean Ṽ(0) = {np.mean(V0_paths):.4f}"
        )

        # (c) Build supervised dataset
        inputs_np, targets_np = _build_targets(Y_jumped, time_grid, X_paths, cfg)

        # (d) Train network
        logger.info(f"  Training Y_net ({cfg.num_train_steps} steps)...")
        losses = _optimise(net, inputs_np, targets_np, cfg)
        all_losses.extend(losses)

    # Final evaluation after last training round
    Y_raw_final    = _evaluate_network(net, time_grid, X_paths, cfg)
    Y_final_jumped = _apply_exercise_jumps(Y_raw_final, U_tilde, ex_step_idx)
    Y_snapshots.append(Y_final_jumped)

    logger.info(
        f"\nTraining complete. "
        f"Final mean Ṽ(0) = {np.mean(Y_final_jumped[:, 0]):.4f}"
    )
    logger.info("Stage 2 complete.\n")
    return net, all_losses, Y_snapshots


# ─────────────────────────────────────────────────────────────────────────────
# 7.  Stage 3 — Exposure & Exercise-Time Artifacts
# ─────────────────────────────────────────────────────────────────────────────

def compute_exposure(
    cfg: Config,
    hw: HullWhite,
    net: nn.Module,
    pre_train_data: Dict,
) -> Dict:
    """
    Stage 3 of She & Grecu (2018) Sections 2.3 and 3.1.

    (a) Re-evaluate trained Y_net with exercise jumps.
    (b) Determine exercise indicators η_{pm} at each exercise date:
            η_{pm} = 0  if path p exercises at T_m,  1 otherwise   (eq 3.2)
        A path exercises when U_tilde(T_m) > continuation value (before jump).
    (c) Build cumulative non-exercise indicator per time step:
            η̃_{pn} = ∏_{m: T_m ≤ T_n} η_{pm}
        η̃_{pn} = 0 means the option was already exercised before T_n.
    (d) Convert Ṽ back to undiscounted V = Ṽ * B, then zero out exercised paths
        (cash-settled: exposure vanishes after exercise, eq 2.18).
    (e) EPE(T_n) = E[max(V_exposure, 0)],  ENE(T_n) = E[min(V_exposure, 0)].
    (f) CVA = (1-R) * Σ_n EPE(T_n) * ΔPD(T_n)  with flat hazard rate λ.
    (g) Exercise time per path τ_p = first T_m where η_{pm} = 0, else NaN.

    Args:
        cfg            : Config
        hw             : HullWhite
        net            : trained ValueNetwork
        pre_train_data : output of pre_training()

    Returns
    -------
    dict with keys:
        'time_grid'      : [N+1]
        'EPE'            : [N+1]    discounted expected positive exposure
        'ENE'            : [N+1]    discounted expected negative exposure
        'V_exposure'     : [M, N+1] per-path exposure (zero after exercise)
        'exercise_times' : [M]      τ_p per path (NaN = unexercised)
        'exercise_prob'  : [num_ex] P(exercised at T_m | alive at T_m)
        'survival_curve' : [num_ex] P(not yet exercised by T_m)
        'CVA'            : scalar
        'eta'            : [M, num_ex] exercise indicators
        'Y_jumped'       : [M, N+1] final jump-corrected discounted values
        'ex_step_idx'    : [num_ex]
    """
    logger.info("=" * 60)
    logger.info("STAGE 3 — Post-training: exposure & exercise time artifacts")

    time_grid   = pre_train_data['time_grid']
    X_paths     = pre_train_data['X_paths']
    U_tilde     = pre_train_data['U_tilde']
    U_exercise  = pre_train_data['U_exercise']
    ex_step_idx = pre_train_data['ex_step_idx']
    B_exercise  = pre_train_data['B_exercise']

    M  = cfg.num_paths
    N  = cfg.N
    dt = cfg.dt

    # ── (a)  Re-evaluate with exercise jumps ──────────────────────────────────
    Y_raw    = _evaluate_network(net, time_grid, X_paths, cfg)
    Y_jumped = _apply_exercise_jumps(Y_raw, U_tilde, ex_step_idx)

    # ── (b)  Exercise indicators η_{pm} ──────────────────────────────────────
    # A path exercises at T_m when the exercise value is positive AND exceeds
    # the raw continuation value (i.e. the jump is binding at that step)
    num_ex = len(cfg.exercise_dates)
    eta    = np.ones((M, num_ex), dtype=float)

    for m_idx, step_m in enumerate(ex_step_idx):
        continuation = Y_raw[:, step_m]           # Ṽ before exercise decision
        exercise_val = U_tilde[:, m_idx]           # Ũ(T_m)
        exercised    = (exercise_val > continuation) & (exercise_val > 0)
        eta[exercised, m_idx] = 0.0

    # ── (c)  Numeraire B(t) along all paths for undiscounting ─────────────────
    B_all = np.zeros((M, N + 1))
    for i in range(N + 1):
        r_path_i = X_paths[:, :i + 1] + cfg.r0
        integral  = np.trapezoid(r_path_i, dx=dt, axis=1)
        B_all[:, i] = np.exp(integral)

    # ── (d)  Undiscounted exposure, zeroed on exercised paths ─────────────────
    # Start from discounted Ṽ, convert to undiscounted V = Ṽ * B
    V_undi = Y_jumped * B_all    # [M, N+1]

    # Build alive mask: 1 if option still open at t_i, 0 if already exercised
    alive_mask = np.ones((M, N + 1))
    for m_idx, step_m in enumerate(ex_step_idx):
        # After T_m: zero out paths that exercised at T_m
        alive_mask[:, step_m + 1:] *= eta[:, m_idx, None]

    # Cash-settled: exposure = 0 on exercised paths (eq 2.18)
    V_exposure = V_undi * alive_mask    # [M, N+1]

    # ── (e)  EPE and ENE (paper eqs 2.5–2.6) ─────────────────────────────────
    EPE = np.mean(np.maximum(V_exposure, 0.0), axis=0)   # [N+1]
    ENE = np.mean(np.minimum(V_exposure, 0.0), axis=0)   # [N+1]

    # ── (f)  CVA with flat hazard rate ────────────────────────────────────────
    # CVA = (1-R) * Σ_n EPE(T_n) * ΔPD(T_n)
    # ΔPD(T_n) = e^{-λ T_{n-1}} - e^{-λ T_n}
    lam      = cfg.hazard_rate
    R        = cfg.recovery
    t_prev   = time_grid[:-1]
    t_curr   = time_grid[1:]
    delta_pd = np.exp(-lam * t_prev) - np.exp(-lam * t_curr)
    CVA      = (1 - R) * float(np.sum(EPE[1:] * delta_pd))
    logger.info(f"  CVA = {CVA:.4f}  (λ={lam}, R={R})")

    # ── (g)  Exercise time per path ───────────────────────────────────────────
    exercise_times = np.full(M, np.nan)
    for m_idx, T_m in enumerate(cfg.exercise_dates):
        newly_exercised = (eta[:, m_idx] == 0.0) & np.isnan(exercise_times)
        exercise_times[newly_exercised] = T_m

    # Exercise probability at each T_m = fraction of paths that exercise there
    exercise_prob = np.mean(eta == 0.0, axis=0)

    # Survival curve: P(not yet exercised by T_m)
    survival_curve = np.array([
        np.mean(np.all(eta[:, :m + 1] == 1.0, axis=1))
        for m in range(num_ex)
    ])

    frac_ex = np.mean(~np.isnan(exercise_times))
    logger.info(f"  Fraction of paths exercised: {frac_ex:.1%}")
    for T_m, p in zip(cfg.exercise_dates, exercise_prob):
        logger.info(f"    T={T_m:.1f}y: exercise prob = {p:.1%}")
    logger.info("Stage 3 complete.\n")

    return {
        'time_grid':      time_grid,
        'EPE':            EPE,
        'ENE':            ENE,
        'V_exposure':     V_exposure,
        'exercise_times': exercise_times,
        'exercise_prob':  exercise_prob,
        'survival_curve': survival_curve,
        'CVA':            CVA,
        'eta':            eta,
        'Y_jumped':       Y_jumped,
        'ex_step_idx':    ex_step_idx,
        'X_paths':        X_paths,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 8.  Bachelier Fit
# ─────────────────────────────────────────────────────────────────────────────

def bachelier_formula(x: np.ndarray, A: float, c: float, s: float) -> np.ndarray:
    """
    Bachelier-type option price (paper eq 3.8):
        V_Bach(x) = A * [(x-c) * Φ((x-c)/s) + s * φ((x-c)/s)]

    Interpretation:
        A : slope of the linear exercise value (DV01-like sensitivity)
        c : the x-value where the exercise value crosses zero (the ATM strike)
        s : effective normal volatility (controls the ATM rounding)

    As x → +∞: V_Bach(x) → A*(x-c)  — approaches the exercise value line.
    As x → -∞: V_Bach(x) → 0         — deep OTM, option worthless.
    """
    z   = (x - c) / (s + 1e-12)
    return A * ((x - c) * scipy_norm.cdf(z) + s * scipy_norm.pdf(z))


def fit_bachelier(
    x_vals: np.ndarray,
    V_vals: np.ndarray,
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Fit the Bachelier formula to a (x, V) scatter at one exercise date.

    Args:
        x_vals : x(T_m) per path, shape [M]
        V_vals : V(T_m) per path (undiscounted), shape [M]

    Returns:
        popt  : [A, c, s]  optimal parameters
        V_fit : [M]        fitted values at x_vals
    """
    # Initial guess: slope from std ratio, ATM at zero, spread from x std
    p0 = [
        np.std(V_vals) / (np.std(x_vals) + 1e-12),  # A
        0.0,                                           # c
        np.std(x_vals),                               # s
    ]
    try:
        popt, _ = curve_fit(
            bachelier_formula, x_vals, V_vals,
            p0=p0, maxfev=10_000,
            bounds=([0, -np.inf, 1e-8], [np.inf, np.inf, np.inf]),
        )
    except RuntimeError:
        popt = np.array(p0)   # fall back to initial guess

    return popt, bachelier_formula(x_vals, *popt)


# ─────────────────────────────────────────────────────────────────────────────
# 9.  Plotting Functions
# ─────────────────────────────────────────────────────────────────────────────

def plot_epe_ene(exposure_data: Dict, cfg: Config):
    """
    Figure 3.1 (left): EPE and ENE of the cash-settled Bermudan swaption.

    Expected shape:
        EPE decreases over time, with downward jumps at each exercise date
        (because exposure vanishes on paths that exercise). ENE ≈ 0 since
        the receiver swaption value is always non-negative.
    """
    fig, ax = plt.subplots(figsize=(9, 5))
    t = exposure_data['time_grid']

    ax.plot(t, exposure_data['EPE'], 'b-', linewidth=1.5, label='EPE')
    ax.plot(t, exposure_data['ENE'], 'r-', linewidth=1.5, label='ENE')

    # Mark exercise dates with vertical dotted lines
    for T_m in cfg.exercise_dates:
        ax.axvline(T_m, color='grey', linestyle=':', alpha=0.5, linewidth=0.8)

    ax.set_xlabel('Time (years)', fontsize=12)
    ax.set_ylabel('Exposure', fontsize=12)
    ax.set_title(
        'EPE and ENE — Cash-Settled Bermudan Swaption\n'
        '(She & Grecu 2018, Fig 3.1 left)',
        fontsize=12,
    )
    ax.legend(fontsize=11)
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    return fig


def plot_loss_curve(all_losses: List[float], cfg: Config):
    """
    Figure 3.1 (right): Training loss vs gradient step.
    Paper shows convergence (plateau) around step 500.
    """
    fig, ax = plt.subplots(figsize=(7, 5))
    ax.semilogy(np.arange(1, len(all_losses) + 1), all_losses,
                'b-', linewidth=0.8, alpha=0.85)
    ax.set_xlabel('Training Step', fontsize=12)
    ax.set_ylabel('MSE Loss (log scale)', fontsize=12)
    ax.set_title(
        'Training Loss Evolution\n(She & Grecu 2018, Fig 3.1 right)',
        fontsize=12,
    )
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    return fig


def plot_future_value_evolution(
    exposure_data: Dict,
    pre_train_data: Dict,
    Y_snapshots: List[np.ndarray],
    exercise_date_idx: int,
    cfg: Config,
):
    """
    Figures 3.2 and 3.3: Scatter of (x, V) at a given exercise date
    at successive Picard iteration snapshots.

    Blue dots  : portfolio value Ṽ(T_m) × B(T_m)  (learned, changes with training)
    Orange dots: exercise value U(T_m)              (fixed, analytical)

    Args:
        exercise_date_idx : 0 → Fig 3.2 (1st exercise date)
                            3 → Fig 3.3 (4th exercise date)
    """
    step_m = pre_train_data['ex_step_idx'][exercise_date_idx]
    T_m    = cfg.exercise_dates[exercise_date_idx]
    x_Tm   = pre_train_data['X_paths'][:, step_m]
    U_Tm   = pre_train_data['U_exercise'][:, exercise_date_idx]
    B_Tm   = pre_train_data['B_exercise'][:, exercise_date_idx]

    n_snaps = len(Y_snapshots)
    n_cols  = min(3, n_snaps)
    n_rows  = math.ceil(n_snaps / n_cols)

    fig, axes = plt.subplots(n_rows, n_cols, figsize=(5 * n_cols, 4 * n_rows))
    axes = np.array(axes).flatten()

    for k, Y_snap in enumerate(Y_snapshots):
        ax    = axes[k]
        V_Tm  = Y_snap[:, step_m] * B_Tm   # convert discounted → undiscounted

        ax.scatter(x_Tm, V_Tm, s=2, alpha=0.3, color='steelblue', label='V (learned)')
        ax.scatter(x_Tm, U_Tm, s=2, alpha=0.3, color='orange',    label='U (exercise)')
        ax.set_title(f'Picard iter {k + 1}', fontsize=10)
        ax.set_xlabel('x', fontsize=9)
        ax.set_ylabel('V', fontsize=9)
        ax.grid(True, alpha=0.2)
        if k == 0:
            ax.legend(fontsize=8, markerscale=4)

    for k in range(n_snaps, len(axes)):
        axes[k].set_visible(False)

    fig_num = '3.2' if exercise_date_idx == 0 else '3.3'
    fig.suptitle(
        f'Future Value Evolution — Exercise Date {exercise_date_idx + 1} (T={T_m:.1f}y)\n'
        f'(She & Grecu 2018, Fig {fig_num})',
        fontsize=12,
    )
    plt.tight_layout()
    return fig


def plot_bachelier_fit(
    exposure_data: Dict,
    pre_train_data: Dict,
    exercise_date_idx: int,
    cfg: Config,
):
    """
    Figure 3.4: Fit of the learned V(x) at an exercise date to the Bachelier
    option price formula (eq 3.8).

    Blue dots    : (x, V) scatter from MC paths
    Red dashed   : fitted Bachelier curve
    """
    step_m = pre_train_data['ex_step_idx'][exercise_date_idx]
    T_m    = cfg.exercise_dates[exercise_date_idx]
    x_Tm   = pre_train_data['X_paths'][:, step_m]
    B_Tm   = pre_train_data['B_exercise'][:, exercise_date_idx]
    V_Tm   = exposure_data['Y_jumped'][:, step_m] * B_Tm   # undiscounted

    popt, _ = fit_bachelier(x_Tm, V_Tm)
    A, c, s = popt

    x_sorted = np.sort(x_Tm)
    V_fitted = bachelier_formula(x_sorted, A, c, s)

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.scatter(x_Tm, V_Tm, s=3, alpha=0.25, color='blue', label='V (learned)')
    ax.plot(x_sorted, V_fitted, 'r--', linewidth=2,
            label=f'Bachelier fit\nA={A:.3f},  c={c:.5f},  s={s:.5f}')

    ax.set_xlabel('x', fontsize=12)
    ax.set_ylabel('V', fontsize=12)
    ax.set_title(
        f'Bachelier Fit — Exercise Date {exercise_date_idx + 1} (T={T_m:.1f}y)\n'
        '(She & Grecu 2018, Fig 3.4)',
        fontsize=12,
    )
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    return fig


def plot_exercise_time_distribution(exposure_data: Dict, cfg: Config):
    """
    Exercise time artifacts (not in paper, additional diagnostic):
        (a) Histogram of exercise times τ_p
        (b) Exercise probability at each T_m (conditional on path reaching T_m)
        (c) Survival curve P(not yet exercised by T_m)
    """
    exercise_times = exposure_data['exercise_times']
    exercise_prob  = exposure_data['exercise_prob']
    survival_curve = exposure_data['survival_curve']
    ex_dates       = cfg.exercise_dates

    fig, axes = plt.subplots(1, 3, figsize=(15, 5))

    # (a) Histogram
    exercised = exercise_times[~np.isnan(exercise_times)]
    axes[0].hist(exercised, bins=len(ex_dates), color='steelblue',
                 edgecolor='black', alpha=0.7)
    axes[0].set_xlabel('Exercise Time (years)', fontsize=11)
    axes[0].set_ylabel('Number of Paths', fontsize=11)
    axes[0].set_title('Exercise Time Distribution', fontsize=12)
    axes[0].grid(True, alpha=0.3)
    axes[0].text(
        0.05, 0.95,
        f'{100 * len(exercised) / len(exercise_times):.1f}% exercised',
        transform=axes[0].transAxes, fontsize=10, va='top',
    )

    # (b) Exercise probability per date
    axes[1].bar(ex_dates, exercise_prob, width=0.3,
                color='orange', edgecolor='black', alpha=0.8)
    axes[1].set_xlabel('Exercise Date (years)', fontsize=11)
    axes[1].set_ylabel('Probability', fontsize=11)
    axes[1].set_title('Exercise Probability by Date', fontsize=12)
    axes[1].set_ylim(0, 1)
    axes[1].grid(True, alpha=0.3)

    # (c) Survival curve
    axes[2].step(ex_dates, survival_curve, where='post',
                 color='green', linewidth=2)
    axes[2].scatter(ex_dates, survival_curve, color='green', s=60, zorder=3)
    axes[2].set_xlabel('Time (years)', fontsize=11)
    axes[2].set_ylabel('P(not yet exercised)', fontsize=11)
    axes[2].set_title('Survival Curve', fontsize=12)
    axes[2].set_ylim(0, 1.05)
    axes[2].grid(True, alpha=0.3)

    plt.suptitle('Exercise Time Artifacts', fontsize=13, fontweight='bold')
    plt.tight_layout()
    return fig


# ─────────────────────────────────────────────────────────────────────────────
# 10.  Excel Export
# ─────────────────────────────────────────────────────────────────────────────

def export_to_excel(
    cfg: Config,
    pre_train_data: Dict,
    exposure_data: Dict,
    all_losses: List[float],
    output_path: str = "bermudan_swaption_results.xlsx",
) -> None:
    """
    Export all results to a formatted 5-sheet Excel workbook.

    Sheets:
        Summary          — model/trade parameters + headline results
        Exposure Profiles— EPE, ENE at every time step
        Exercise Times   — per-path τ_p + exercised flag
        Exercise Analytics— probability by date, survival curve
        Training Loss    — MSE loss at every gradient step

    Formatting:
        Dark-blue headers (white text), alternating light-blue row fills,
        thin borders, auto-width columns, 4 decimal places on floats.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    t_grid         = exposure_data['time_grid']
    ex_times       = exposure_data['exercise_times']
    exercise_prob  = exposure_data['exercise_prob']
    survival_curve = exposure_data['survival_curve']

    # ── DataFrames ────────────────────────────────────────────────────────────
    df_exposure = pd.DataFrame({
        'Time (years)':    t_grid,
        'EPE':             exposure_data['EPE'],
        'ENE':             exposure_data['ENE'],
        'Discount Factor': np.exp(-cfg.r0 * t_grid),
    })

    df_exercise = pd.DataFrame({
        'Path':          np.arange(cfg.num_paths),
        'Exercise Time': ex_times,
        'Exercised':     (~np.isnan(ex_times)).astype(int),
    })

    df_analytics = pd.DataFrame({
        'Exercise Date (years)': cfg.exercise_dates,
        'Exercise Probability':  exercise_prob,
        'Survival Probability':  survival_curve,
    })

    df_loss = pd.DataFrame({
        'Training Step': np.arange(1, len(all_losses) + 1),
        'MSE Loss':      all_losses,
    })

    # ── Write sheets ──────────────────────────────────────────────────────────
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_exposure.to_excel(writer,  sheet_name='Exposure Profiles',  index=False)
        df_exercise.to_excel(writer,  sheet_name='Exercise Times',      index=False)
        df_analytics.to_excel(writer, sheet_name='Exercise Analytics',  index=False)
        df_loss.to_excel(writer,      sheet_name='Training Loss',        index=False)

        wb = writer.book

        # Shared styles
        hdr_font   = Font(bold=True, color='FFFFFF')
        hdr_fill   = PatternFill('solid', fgColor='1F4E79')
        alt_fill   = PatternFill('solid', fgColor='EBF3FB')
        ctr_align  = Alignment(horizontal='center')
        thin       = Side(style='thin')
        brd        = Border(left=thin, right=thin, top=thin, bottom=thin)

        for sname in wb.sheetnames:
            ws = wb[sname]
            # Header
            for cell in ws[1]:
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.alignment = ctr_align
                cell.border    = brd
            # Data rows
            for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                fill = alt_fill if r_idx % 2 == 0 else PatternFill()
                for cell in row:
                    cell.fill   = fill
                    cell.border = brd
                    if isinstance(cell.value, float):
                        cell.number_format = '0.0000'
            # Column widths
            for col in ws.columns:
                max_w = max(
                    len(str(c.value)) if c.value is not None else 0
                    for c in col
                )
                ws.column_dimensions[get_column_letter(col[0].column)].width = (
                    min(max_w + 4, 28)
                )

        # Summary sheet (inserted at position 0)
        ws_sum = wb.create_sheet('Summary', 0)
        summary_rows = [
            ('Parameter',            'Value'),
            ('Notional',              cfg.notional),
            ('Fixed Rate',            f'{cfg.fixed_rate:.2%}'),
            ('Swap Tenor (years)',     cfg.swap_tenor),
            ('Exercise Start',         cfg.exercise_start),
            ('Exercise End',           cfg.exercise_end),
            ('Exercise Frequency',     cfg.exercise_freq),
            ('kappa (κ)',              cfg.kappa),
            ('sigma_r (σ_r)',          cfg.sigma_r),
            ('r0 (flat forward rate)', f'{cfg.r0:.2%}'),
            ('Num Paths',              cfg.num_paths),
            ('Num Time Steps (N)',      cfg.N),
            ('dt',                     f'{cfg.dt:.5f}'),
            ('Hazard Rate (λ)',         f'{cfg.hazard_rate:.2%}'),
            ('Recovery Rate (R)',       f'{cfg.recovery:.0%}'),
            ('─' * 20,                 '─' * 20),
            ('CVA',                    f'{exposure_data["CVA"]:.6f}'),
            ('EPE at t=0',             f'{exposure_data["EPE"][0]:.4f}'),
            ('% Paths Exercised',
             f'{100 * np.mean(~np.isnan(exposure_data["exercise_times"])):.1f}%'),
            ('Final Training Loss',    f'{all_losses[-1]:.4e}'),
        ]
        for row in summary_rows:
            ws_sum.append(row)
        for cell in ws_sum[1]:
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = ctr_align
        for col in ws_sum.columns:
            ws_sum.column_dimensions[get_column_letter(col[0].column)].width = 30

    logger.info(f"Results exported → {output_path}")
