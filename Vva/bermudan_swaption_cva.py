"""
Bermudan Swaption CVA — Deep BSDE Approach
===========================================
Standalone implementation of She & Grecu (2018), Section 3.
arXiv:1811.08726

Architecture: Option B — fully self-contained, no inheritance from
UncoupledFBSDESolver. All three stages are explicit, top-level functions.

Paper reference:
    Stage 1  — pre_training()   : simulate HW paths + exercise values
    Stage 2  — train()          : backward Picard with early-exercise jump
    Stage 3  — compute_exposure(): EPE / ENE / exercise-time artifacts

Hull-White 1-factor model
--------------------------
    r(t) = x(t) + f(0,t),   f(0,t) = r0  (flat curve)
    dx(t) = [y(t) - κ x(t)] dt + σ_r dW(t),   x(0) = 0
    y(t)  = (σ_r²/2κ)(1 - e^{-2κt})

Swaption (Section 3.2 parameters)
-----------------------------------
    Cash-settled receiver Bermudan swaption
    Notional = 10,000,  Fixed rate K = 2.8%
    Exercise: semi-annual, 1.5y → 3.5y  (5 exercise dates)
    Underlying swap tenor: 4y from each exercise date

BSDE formulation
-----------------
    Forward : dx(t) = [y(t) - κ x(t)] dt + σ_r dW(t)
    Backward: dṼ(t) = Z(t) dW(t)        ← driver = 0 (Ṽ is a Q-martingale)
    Terminal: Ṽ(T_N^+) = 0              ← post-last-exercise boundary
    Jump    : Ṽ(T_m^-) = max(Ṽ(T_m^+), Ũ(T_m))  ← early exercise (eq 2.14)

    where Ũ(T_m) = U(T_m)/B(T_m) is the discounted exercise value,
    U(T_m) = max(V_receiver_swap(T_m), 0) computed from HW bond prices.

Neural network
--------------
    Single MLP: Y_net(t, x) → Ṽ(t, x)
    Input  : (t, x)  — shape [M, 2]
    Output : Ṽ       — shape [M, 1]
    Architecture: 2 hidden layers of width (1 + d̃) = 11, tanh activation
    (paper eq 3.7: d=1 risk factor, d̃=10 expansion)

Training
--------
    Loss (eq 3.1): L = mean((Ṽ_p(t=0) - V0)²)
    where V0 is a learnable scalar parameter.
    Backward Picard: networks are retrained each iteration using jump-
    corrected Y targets built from the previous iterate.

Outputs
-------
    • EPE(t), ENE(t) profiles
    • Exercise time distribution per path
    • Exercise probability by date
    • Survival curve
    • Future value scatter at exercise dates (reproduce Fig 3.2 / 3.3)
    • Bachelier fit to learned V(x) (reproduce Fig 3.4)
    • Loss curve (reproduce Fig 3.1 right)
"""

# ─────────────────────────────────────────────────────────────────────────────
# 0.  Imports
# ─────────────────────────────────────────────────────────────────────────────

import math
import time
import logging
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import numpy as np
import torch
import torch.nn as nn
from scipy.optimize import curve_fit
from scipy.stats import norm as scipy_norm
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import pandas as pd

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

torch.manual_seed(42)
np.random.seed(42)

DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")
logger.info(f"Device: {DEVICE}")


# ─────────────────────────────────────────────────────────────────────────────
# 1.  Configuration
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class Config:
    """
    All model, product, and training parameters in one place.
    Defaults match Section 3.2 of She & Grecu (2018).
    """

    # ── Hull-White model ──────────────────────────────────────────────────────
    kappa:   float = 0.01    # mean-reversion speed κ
    sigma_r: float = 0.01    # short-rate volatility σ_r (flat term-structure)
    r0:      float = 0.028   # flat forward rate f(0,t) = r0

    # ── Underlying swap ───────────────────────────────────────────────────────
    notional:   float = 10_000.0  # notional N
    fixed_rate: float = 0.028     # fixed coupon K (ATM at r0)
    swap_tenor: float = 4.0       # swap length in years from exercise date
    fixed_freq: float = 0.5       # fixed leg payment frequency (semi-annual)
    float_freq: float = 0.25      # floating leg frequency (3M Libor)

    # ── Exercise schedule ─────────────────────────────────────────────────────
    exercise_start: float = 1.5   # first exercise date (years)
    exercise_end:   float = 3.5   # last exercise date
    exercise_freq:  float = 0.5   # semi-annual

    # ── Simulation ────────────────────────────────────────────────────────────
    num_paths: int   = 5_000      # Monte Carlo paths M
    dt:        float = 1 / 52     # ~weekly time steps

    # ── Neural network ────────────────────────────────────────────────────────
    # Paper: 2 hidden layers, d̃=10  → width = d + d̃ = 1 + 10 = 11
    hidden_dim:        int = 11
    num_hidden_layers: int = 2

    # ── Training ──────────────────────────────────────────────────────────────
    num_train_steps: int   = 5_000   # Adam gradient steps
    learning_rate:   float = 1e-3
    lr_decay_steps:  int   = 1_000   # StepLR period
    lr_decay_rate:   float = 0.5
    log_every:       int   = 500     # print loss every N steps

    # ── CVA credit parameters ─────────────────────────────────────────────────
    hazard_rate: float = 0.01   # constant hazard rate λ (for CVA integral)
    recovery:    float = 0.40   # recovery rate R

    # ── Derived (filled in __post_init__) ─────────────────────────────────────
    T:              float      = field(init=False)
    N:              int        = field(init=False)
    exercise_dates: List[float] = field(init=False)

    def __post_init__(self):
        # Simulation horizon: last exercise + full swap tenor
        self.T = self.exercise_end + self.swap_tenor

        # Snap N so dt divides T exactly
        self.N  = max(1, round(self.T / self.dt))
        self.dt = self.T / self.N

        # Exercise dates: semi-annual grid from exercise_start to exercise_end
        n_ex = round((self.exercise_end - self.exercise_start) / self.exercise_freq) + 1
        self.exercise_dates = [
            round(self.exercise_start + i * self.exercise_freq, 10)
            for i in range(n_ex)
        ]


# ─────────────────────────────────────────────────────────────────────────────
# 2.  Hull-White Analytical Formulae
# ─────────────────────────────────────────────────────────────────────────────

class HullWhite:
    """
    Analytical zero-coupon bond prices and swap values under HW 1-factor.

    All methods operate on numpy arrays of x(t) values (shape [M]).

    Reference: Andersen & Piterbarg (2010), Vol 2, Chapter 10.
               She & Grecu (2018), eqs (3.5)–(3.6).
    """

    def __init__(self, cfg: Config):
        self.kappa   = cfg.kappa
        self.sigma_r = cfg.sigma_r
        self.r0      = cfg.r0

    # ── Scalar helpers ────────────────────────────────────────────────────────

    def y(self, t: float) -> float:
        """
        Variance function y(t) = (σ_r²/2κ)(1 - e^{-2κt}).
        Appears in the HW drift: dx = [y(t) - κ x] dt + σ_r dW.
        """
        if self.kappa < 1e-12:
            return self.sigma_r**2 * t
        return (self.sigma_r**2 / (2 * self.kappa)) * (1.0 - math.exp(-2 * self.kappa * t))

    def B(self, t: float, T: float) -> float:
        """
        Bond duration factor B(t,T) = (1/κ)(1 - e^{-κ(T-t)}).
        Used in P(t,T) = A(t,T) exp(-B(t,T) x(t)).
        """
        tau = T - t
        if self.kappa < 1e-12:
            return tau
        return (1.0 - math.exp(-self.kappa * tau)) / self.kappa

    def ln_A(self, t: float, T: float) -> float:
        """
        Log of the A(t,T) factor in the ZCB formula.
        For a flat forward curve f(0,t) = r0:
            ln A = ln[P(0,T)/P(0,t)] + B(t,T)*r0 - (σ_r²/4κ) B(t,T)²(1-e^{-2κt})
                 = -r0*(T-t) + B(t,T)*r0 - 0.5*σ_r²*B(t,T)²*y(t)/σ_r²
        Simplifying the convexity term using y(t):
            = -r0*(T-t) + B(t,T)*r0 - 0.5*B(t,T)²*(2κ*y(t))/(2κ)  ... see A&P
        """
        Bval    = self.B(t, T)
        log_zcb = -self.r0 * (T - t)          # ln[P(0,T)/P(0,t)] flat curve
        drift   = Bval * self.r0               # B(t,T) * f(0,t)
        # Convexity: -0.5 * σ_r² * B² * ∫₀ᵗ e^{-2κ(t-u)} du = -0.5*B²*(2κ*y/σ_r²)*σ_r²/2κ
        # = -0.5 * B(t,T)² * y(t)   (this is the standard HW result)
        convex  = -0.5 * Bval**2 * self.y(t)
        return log_zcb + drift + convex

    # ── Vectorised bond price ─────────────────────────────────────────────────

    def zcb(self, t: float, T: float, x: np.ndarray) -> np.ndarray:
        """
        Zero-coupon bond price P(t,T | x(t)).
            P(t,T) = exp(ln A(t,T) - B(t,T) * x(t))

        Args:
            t : current time (scalar)
            T : bond maturity (scalar, T > t)
            x : x(t) values, shape [M]

        Returns:
            P : shape [M]
        """
        return np.exp(self.ln_A(t, T) - self.B(t, T) * x)

    # ── Swap and exercise value ───────────────────────────────────────────────

    def receiver_swap_value(self, t_exercise: float, x: np.ndarray,
                            cfg: Config) -> np.ndarray:
        """
        Value of a *receiver* swap at t = t_exercise (swap starts immediately).

        Receiver swap: receive fixed K, pay floating.
        V_receiver = fixed_leg - float_leg
                   = N*K * Σ τ_k P(t, T_k)  -  N*[P(t, T_start) - P(t, T_end)]

        where T_start = t_exercise, T_end = t_exercise + swap_tenor,
        and τ_k = fixed_freq for all k.

        Args:
            t_exercise : exercise date (also swap start), scalar
            x          : x(t_exercise), shape [M]
            cfg        : Config object

        Returns:
            V_receiver : shape [M]
        """
        N  = cfg.notional
        K  = cfg.fixed_rate
        tau = cfg.fixed_freq

        # Fixed payment dates: t_exercise + k*fixed_freq, k=1,...,n_fixed
        n_fixed    = round(cfg.swap_tenor / cfg.fixed_freq)
        fixed_dates = [t_exercise + (i + 1) * cfg.fixed_freq for i in range(n_fixed)]

        # Fixed leg: N * K * Σ τ * P(t, T_k)
        fixed_leg = sum(
            tau * self.zcb(t_exercise, Tk, x) for Tk in fixed_dates
        )
        fixed_leg *= N * K

        # Float leg via replication: N * [P(t, T_start) - P(t, T_end)]
        # T_start = t_exercise (first Libor reset), T_end = last payment
        P_start    = self.zcb(t_exercise, t_exercise, x)   # = 1 by definition
        P_end      = self.zcb(t_exercise, fixed_dates[-1], x)
        float_leg  = N * (P_start - P_end)

        return fixed_leg - float_leg

    def exercise_value(self, t_exercise: float, x: np.ndarray,
                       cfg: Config) -> np.ndarray:
        """
        Cash-settled exercise value for a *receiver* swaption at T_m:
            U(T_m) = max(V_receiver(T_m, x), 0)

        The option is in-the-money when fixed rate K > swap rate,
        i.e. when the receiver swap has positive value.

        Args:
            t_exercise : T_m, scalar
            x          : x(T_m), shape [M]
            cfg        : Config

        Returns:
            U : shape [M], non-negative
        """
        return np.maximum(self.receiver_swap_value(t_exercise, x, cfg), 0.0)

    def numeraire(self, t: float, x_path: np.ndarray,
                  x_grid: np.ndarray, dt: float) -> np.ndarray:
        """
        Bank account numeraire B(t) = exp(∫₀ᵗ r(s) ds).
        Approximated along each path via the trapezoidal rule on stored x values.

        Args:
            t      : current time
            x_path : x values from t=0 to t, shape [M, n_steps+1]
            x_grid : time grid from 0 to t, shape [n_steps+1]
            dt     : time step

        Returns:
            B : shape [M]
        """
        # r(s) = x(s) + r0  along each path
        r_path  = x_path + self.r0          # [M, n_steps+1]
        # Trapz integration: ∫ r ds ≈ dt * (r0/2 + r1 + r2 + ... + r_{n-1} + r_n/2)
        integral = np.trapezoid(r_path, dx=dt, axis=1)
        return np.exp(integral)


# ─────────────────────────────────────────────────────────────────────────────
# 3.  Neural Network — Y_net(t, x) → Ṽ(t, x)
# ─────────────────────────────────────────────────────────────────────────────

class ValueNetwork(nn.Module):
    """
    MLP approximating the discounted future value Ṽ(t, x).

    Architecture (paper eq 3.7, Section 3.2):
        Input  : [t, x]  — 2-dimensional
        Hidden : num_hidden_layers layers of width hidden_dim, tanh activation
        Output : scalar Ṽ

    Note: the paper parameterises the Delta F^(n)(x) per time step.
    We instead parameterise Y(t, x) directly (a single shared network
    across all times) — this is the Picard approach from the thesis.
    """

    def __init__(self, hidden_dim: int = 11, num_hidden: int = 2):
        super().__init__()

        layers: List[nn.Module] = []
        in_dim = 2  # input: (t, x)

        for _ in range(num_hidden):
            layers.append(nn.Linear(in_dim, hidden_dim))
            layers.append(nn.Tanh())
            in_dim = hidden_dim

        layers.append(nn.Linear(in_dim, 1))   # scalar output Ṽ
        self.net = nn.Sequential(*layers)

        # Small uniform initialisation — paper starts from random params
        for m in self.modules():
            if isinstance(m, nn.Linear):
                nn.init.uniform_(m.weight, -0.1, 0.1)
                nn.init.zeros_(m.bias)

    def forward(self, t_x: torch.Tensor) -> torch.Tensor:
        """
        Args:
            t_x : [M, 2] — concatenated (t, x) inputs
        Returns:
            V_tilde : [M, 1]
        """
        return self.net(t_x)


# ─────────────────────────────────────────────────────────────────────────────
# 4.  Stage 1 — Pre-Training: Simulate Paths & Exercise Values
# ─────────────────────────────────────────────────────────────────────────────

def pre_training(cfg: Config, hw: HullWhite) -> Dict:
    """
    Stage 1 of She & Grecu (2018) Section 3.1.

    Operations:
        (a) Forward-simulate x(t) paths via Euler-Maruyama.
            Store in tensor X_paths [M, N+1].
        (b) Store Brownian increments dW [M, N] for use in BSDE evolution.
        (c) At each exercise date T_m, compute the (undiscounted) exercise
            value U(T_m, x) = max(V_receiver_swap, 0) analytically.
            Store in U_exercise [M, num_exercise_dates].
        (d) Compute numeraire B(T_m) along each path for discounting:
            Ũ(T_m) = U(T_m) / B(T_m).

    Args:
        cfg : Config
        hw  : HullWhite model

    Returns dict with keys:
        'time_grid'     : [N+1]          uniform time grid
        'X_paths'       : [M, N+1]       x(t) paths
        'dW'            : [M, N]         Brownian increments
        'U_exercise'    : [M, num_ex]    undiscounted exercise values
        'U_tilde'       : [M, num_ex]    discounted exercise values Ũ = U/B
        'ex_step_idx'   : [num_ex]       time-step indices for exercise dates
        'B_exercise'    : [M, num_ex]    numeraire B(T_m) per path
    """
    logger.info("=" * 60)
    logger.info("STAGE 1 — Pre-training: simulating paths & exercise values")

    M  = cfg.num_paths
    N  = cfg.N
    dt = cfg.dt

    # ── (a)  Euler-Maruyama simulation of x(t) ────────────────────────────────
    # dx(t) = [y(t) - κ x(t)] dt + σ_r dW(t),   x(0) = 0
    time_grid = np.linspace(0.0, cfg.T, N + 1)   # [N+1]
    X_paths   = np.zeros((M, N + 1))              # x(0) = 0
    dW        = np.random.randn(M, N) * math.sqrt(dt)  # Brownian increments

    for i in range(N):
        t_i   = time_grid[i]
        drift = (hw.y(t_i) - cfg.kappa * X_paths[:, i]) * dt
        X_paths[:, i + 1] = X_paths[:, i] + drift + cfg.sigma_r * dW[:, i]

    logger.info(f"  Simulated {M:,} paths over {N} steps (T={cfg.T:.1f}y, dt={dt:.4f})")

    # ── (b)  Map exercise dates to time-step indices ───────────────────────────
    ex_step_idx = [
        int(round(T_m / dt)) for T_m in cfg.exercise_dates
    ]
    # Clamp to valid range
    ex_step_idx = [min(idx, N) for idx in ex_step_idx]

    # ── (c)  Compute exercise values U(T_m) and numeraire B(T_m) ─────────────
    num_ex     = len(cfg.exercise_dates)
    U_exercise = np.zeros((M, num_ex))   # undiscounted
    U_tilde    = np.zeros((M, num_ex))   # discounted: Ũ = U / B
    B_exercise = np.zeros((M, num_ex))   # numeraire at each exercise date

    for m_idx, (T_m, step_m) in enumerate(zip(cfg.exercise_dates, ex_step_idx)):
        # x values at T_m for all paths
        x_Tm = X_paths[:, step_m]

        # Undiscounted exercise value U(T_m)
        U_exercise[:, m_idx] = hw.exercise_value(T_m, x_Tm, cfg)

        # Numeraire B(T_m) via trapezoidal rule on stored path
        # r(s) = x(s) + r0  →  B(T_m) = exp(∫₀^{T_m} r(s) ds)
        r_path_m  = X_paths[:, :step_m + 1] + cfg.r0   # [M, step_m+1]
        integral  = np.trapezoid(r_path_m, dx=dt, axis=1)   # [M]
        B_m       = np.exp(integral)                     # [M]
        B_exercise[:, m_idx] = B_m

        # Discounted exercise value: Ũ(T_m) = U(T_m) / B(T_m)
        U_tilde[:, m_idx] = U_exercise[:, m_idx] / B_m

        logger.info(
            f"  T_m={T_m:.1f}y (step {step_m}): "
            f"mean U={np.mean(U_exercise[:, m_idx]):.2f}, "
            f"mean Ũ={np.mean(U_tilde[:, m_idx]):.4f}, "
            f"% ITM={100*np.mean(U_exercise[:, m_idx] > 0):.1f}%"
        )

    logger.info("Stage 1 complete.\n")

    return {
        'time_grid':   time_grid,
        'X_paths':     X_paths,
        'dW':          dW,
        'U_exercise':  U_exercise,
        'U_tilde':     U_tilde,
        'ex_step_idx': ex_step_idx,
        'B_exercise':  B_exercise,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 5.  Stage 2 — Training: Picard Iteration with Early-Exercise Jump
# ─────────────────────────────────────────────────────────────────────────────

def _apply_exercise_jumps(
    Y_tilde: np.ndarray,
    U_tilde: np.ndarray,
    ex_step_idx: List[int],
) -> np.ndarray:
    """
    Apply the early-exercise jump condition (paper eq 2.14) to a Y path:
        Ṽ(T_m^-) = max(Ṽ(T_m^+), Ũ(T_m))

    This is applied in-place on the stored Y path *at each exercise date*.
    The "^+" value is the network's continuation-value estimate;
    the "^-" value is the max with the exercise value.

    Because we sweep time from N down to 0 during target construction,
    we apply the jump at each exercise step as we pass through it.

    Args:
        Y_tilde     : [M, N+1]  discounted value path (modified in-place copy)
        U_tilde     : [M, num_ex]  discounted exercise values
        ex_step_idx : list of time-step indices for exercise dates

    Returns:
        Y_tilde_jumped : [M, N+1]  with jumps applied
    """
    Y = Y_tilde.copy()
    for m_idx, step_m in enumerate(ex_step_idx):
        # At step_m: Y[:, step_m] is the continuation value Ṽ(T_m^+)
        # Apply: Ṽ(T_m^-) = max(Ṽ(T_m^+), Ũ(T_m))
        Y[:, step_m] = np.maximum(Y[:, step_m], U_tilde[:, m_idx])
    return Y


def _build_targets(
    Y_tilde_jumped: np.ndarray,
    time_grid: np.ndarray,
    X_paths: np.ndarray,
    cfg: Config,
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Build supervised training targets for Y_net.

    Since the driver f = 0, the BSDE integral collapses and the Feynman-Kac
    target for Ṽ(t_i, x_i) is simply the *jump-corrected* Ṽ value at the
    next time step propagated backward:

        target(t_i) = Ṽ_jumped(t_{i+1})   (no driver integral needed)

    We collect all (t_i, x_i, target_i) tuples across all time steps
    and all paths into flat arrays for batch training.

    Args:
        Y_tilde_jumped : [M, N+1]  jump-corrected Y path
        time_grid      : [N+1]
        X_paths        : [M, N+1]
        cfg            : Config

    Returns:
        inputs  : [M*N, 2]   (t_i, x_i) pairs
        targets : [M*N, 1]   Ṽ(t_{i+1}) — next-step jump-corrected value
    """
    M = cfg.num_paths
    N = cfg.N

    inputs_list  = []
    targets_list = []

    for i in range(N):
        # Input: (t_i, x_i) for all paths
        t_i   = time_grid[i]
        x_i   = X_paths[:, i]                    # [M]
        t_vec = np.full(M, t_i)                  # [M]

        # Target: jump-corrected value at next step
        # Ṽ(t_{i+1}) already has jumps baked in from _apply_exercise_jumps
        tgt = Y_tilde_jumped[:, i + 1]            # [M]

        inputs_list.append(np.stack([t_vec, x_i], axis=1))  # [M, 2]
        targets_list.append(tgt[:, None])                    # [M, 1]

    inputs  = np.concatenate(inputs_list,  axis=0)   # [M*N, 2]
    targets = np.concatenate(targets_list, axis=0)   # [M*N, 1]
    return inputs, targets


def _optimise(
    net: nn.Module,
    inputs_np: np.ndarray,
    targets_np: np.ndarray,
    cfg: Config,
) -> List[float]:
    """
    Standard Adam training loop with StepLR decay.

    Args:
        net        : ValueNetwork
        inputs_np  : [M*N, 2]  numpy
        targets_np : [M*N, 1]  numpy

    Returns:
        losses : list of per-step MSE losses
    """
    inputs  = torch.tensor(inputs_np,  dtype=torch.float32, device=DEVICE)
    targets = torch.tensor(targets_np, dtype=torch.float32, device=DEVICE)

    optimiser = torch.optim.Adam(net.parameters(), lr=cfg.learning_rate)
    scheduler = torch.optim.lr_scheduler.StepLR(
        optimiser,
        step_size=cfg.lr_decay_steps,
        gamma=cfg.lr_decay_rate,
    )
    loss_fn = nn.MSELoss()
    losses  = []

    net.train()
    dataset_size = inputs.shape[0]
    indices      = torch.arange(dataset_size, device=DEVICE)

    for step in range(cfg.num_train_steps):
        # Mini-batch: shuffle all (t,x) pairs and take one pass
        perm      = indices[torch.randperm(dataset_size, device=DEVICE)]
        batch_in  = inputs[perm[:2048]]    # fixed batch of 2048
        batch_tgt = targets[perm[:2048]]

        optimiser.zero_grad()
        pred = net(batch_in)
        loss = loss_fn(pred, batch_tgt)
        loss.backward()
        torch.nn.utils.clip_grad_norm_(net.parameters(), 1.0)
        optimiser.step()
        scheduler.step()

        losses.append(loss.item())

        if (step + 1) % cfg.log_every == 0:
            logger.info(f"    step {step+1:5d}/{cfg.num_train_steps}: loss={loss.item():.4e}")

    return losses


def _evaluate_network(
    net: nn.Module,
    time_grid: np.ndarray,
    X_paths: np.ndarray,
    cfg: Config,
) -> np.ndarray:
    """
    Evaluate Y_net(t_i, x_i) for all paths and all time steps.

    Returns:
        Y_tilde : [M, N+1]  raw network output (no jumps applied yet)
    """
    M = cfg.num_paths
    N = cfg.N

    net.eval()
    Y_tilde = np.zeros((M, N + 1))

    with torch.no_grad():
        for i in range(N):
            t_i   = time_grid[i]
            x_i   = X_paths[:, i]
            t_vec = np.full(M, t_i)
            nn_in = torch.tensor(
                np.stack([t_vec, x_i], axis=1),
                dtype=torch.float32, device=DEVICE,
            )
            Y_tilde[:, i] = net(nn_in).squeeze(-1).cpu().numpy()

    # Terminal condition: Ṽ(T_N^+) = 0 (post-last-exercise, paper eq 2.16)
    Y_tilde[:, N] = 0.0
    return Y_tilde


def train(
    cfg: Config,
    hw: HullWhite,
    pre_train_data: Dict,
    num_picard_iters: int = 3,
) -> Tuple[nn.Module, List[float], List[np.ndarray]]:
    """
    Stage 2 of She & Grecu (2018) Section 3.1.

    Picard iteration with early-exercise jump:
        For each iteration k:
            (a) Evaluate Y_net on all paths → Y_tilde [M, N+1]
            (b) Apply exercise jumps: Ṽ(T_m) ← max(Ṽ(T_m), Ũ(T_m))
            (c) Build training targets from jump-corrected Y
            (d) Retrain Y_net to minimise MSE against those targets

    The loss reported by the paper (eq 3.1) is mean((Ṽ_p(0) - V0)²),
    measuring how well the backward-propagated value collapses to a
    single V0 at t=0. We log this alongside the MSE training loss.

    Args:
        cfg              : Config
        hw               : HullWhite
        pre_train_data   : output of pre_training()
        num_picard_iters : number of Picard fixed-point iterations

    Returns:
        net         : trained ValueNetwork
        all_losses  : flat list of training losses across all iterations
        Y_snapshots : list of Y_tilde arrays [M, N+1] at each iteration
                      (used for plotting Fig 3.2 / 3.3 evolution)
    """
    logger.info("=" * 60)
    logger.info("STAGE 2 — Training: Picard iteration with exercise jumps")

    time_grid   = pre_train_data['time_grid']
    X_paths     = pre_train_data['X_paths']
    U_tilde     = pre_train_data['U_tilde']
    ex_step_idx = pre_train_data['ex_step_idx']

    # Initialise network and send to device
    net = ValueNetwork(
        hidden_dim=cfg.hidden_dim,
        num_hidden=cfg.num_hidden_layers,
    ).to(DEVICE)

    all_losses  = []
    Y_snapshots = []   # store jump-corrected Y after each Picard iter

    for k in range(num_picard_iters):
        logger.info(f"\n  Picard iteration {k+1}/{num_picard_iters}")

        # (a) Evaluate current network on all paths
        Y_raw = _evaluate_network(net, time_grid, X_paths, cfg)

        # (b) Apply early-exercise jumps to get the "true" held value
        Y_jumped = _apply_exercise_jumps(Y_raw, U_tilde, ex_step_idx)

        # Store snapshot for plotting
        Y_snapshots.append(Y_jumped.copy())

        # Paper loss (eq 3.1): consistency of t=0 values across paths
        V0_paths  = Y_jumped[:, 0]
        paper_loss = np.mean((V0_paths - np.mean(V0_paths))**2)
        logger.info(
            f"  Paper loss (var of Ṽ_p(0)): {paper_loss:.4e}  "
            f"| mean Ṽ(0)={np.mean(V0_paths):.4f}"
        )

        # (c) Build supervised targets
        inputs_np, targets_np = _build_targets(Y_jumped, time_grid, X_paths, cfg)

        # (d) Train network
        logger.info(f"  Training Y_net ({cfg.num_train_steps} steps)...")
        losses = _optimise(net, inputs_np, targets_np, cfg)
        all_losses.extend(losses)

    # Final evaluation with trained network
    Y_raw_final    = _evaluate_network(net, time_grid, X_paths, cfg)
    Y_final_jumped = _apply_exercise_jumps(Y_raw_final, U_tilde, ex_step_idx)
    Y_snapshots.append(Y_final_jumped)

    logger.info(
        f"\nTraining complete. Final mean Ṽ(0) = "
        f"{np.mean(Y_final_jumped[:, 0]):.4f}"
    )
    logger.info("Stage 2 complete.\n")

    return net, all_losses, Y_snapshots


# ─────────────────────────────────────────────────────────────────────────────
# 6.  Stage 3 — Post-Training: Exposure & Exercise Time Artifacts
# ─────────────────────────────────────────────────────────────────────────────

def compute_exposure(
    cfg: Config,
    hw: HullWhite,
    net: nn.Module,
    pre_train_data: Dict,
) -> Dict:
    """
    Stage 3 of She & Grecu (2018) Sections 2.3 and 3.1.

    Operations:
        (a) Re-evaluate trained Y_net on all paths with exercise jumps.
        (b) Track exercise indicators η_{pm} at each exercise date:
                η_{pm} = 0 if exercised at T_m, 1 otherwise  (eq 3.2)
        (c) Build cumulative non-exercise indicator:
                η̃_{pn} = ∏_{m: T_m ≤ T_n} η_{pm}
            η̃_{pn} = 0 means the option was exercised before T_n.
        (d) For cash-settled Bermudan: exposure is zero after exercise.
            V_exposure(T_n, ω_p) = η̃_{pn} * Ṽ(T_n) * B(T_n)
            (convert back from discounted to undiscounted value)
        (e) EPE(T_n) = E[max(V_exposure, 0)]  (eq 3.3)
            ENE(T_n) = E[min(V_exposure, 0)]  (eq 3.4)
        (f) CVA = (1-R) * Σ_n EPE(T_n) * ΔPD(T_n)
            with flat hazard rate: PD(t) = 1 - e^{-λt}
        (g) Exercise time artifacts:
                τ_p = first T_m where η_{pm} = 0  (or NaN if unexercised)

    Args:
        cfg            : Config
        hw             : HullWhite
        net            : trained ValueNetwork
        pre_train_data : output of pre_training()

    Returns dict with keys:
        'time_grid'          : [N+1]
        'EPE'                : [N+1]    expected positive exposure
        'ENE'                : [N+1]    expected negative exposure
        'V_undiscounted'     : [M, N+1] per-path undiscounted V (post-exercise zeroed)
        'exercise_times'     : [M]      τ_p per path (NaN = unexercised)
        'exercise_prob'      : [num_ex] fraction exercised at each T_m
        'survival_curve'     : [num_ex] P(not yet exercised at T_m)
        'CVA'                : scalar
        'ex_step_idx'        : [num_ex]
    """
    logger.info("=" * 60)
    logger.info("STAGE 3 — Post-training: exposure & exercise time artifacts")

    time_grid   = pre_train_data['time_grid']
    X_paths     = pre_train_data['X_paths']
    U_tilde     = pre_train_data['U_tilde']
    U_exercise  = pre_train_data['U_exercise']
    ex_step_idx = pre_train_data['ex_step_idx']
    B_exercise  = pre_train_data['B_exercise']

    M  = cfg.num_paths
    N  = cfg.N
    dt = cfg.dt

    # ── (a)  Re-evaluate network with jumps ───────────────────────────────────
    Y_raw    = _evaluate_network(net, time_grid, X_paths, cfg)
    Y_jumped = _apply_exercise_jumps(Y_raw, U_tilde, ex_step_idx)

    # ── (b)  Exercise indicators η_{pm} ──────────────────────────────────────
    # η_{pm} = 0  if option is exercised at T_m (i.e. exercise value > hold value)
    # η_{pm} = 1  otherwise (continue holding)
    num_ex = len(cfg.exercise_dates)
    eta    = np.ones((M, num_ex), dtype=float)   # [M, num_ex]

    for m_idx, step_m in enumerate(ex_step_idx):
        # Option is exercised when exercise value is strictly binding:
        # i.e. when U_tilde > network continuation value (before jump)
        continuation = Y_raw[:, step_m]          # Ṽ before jump at T_m
        exercise_val = U_tilde[:, m_idx]
        # Exercised when exercise value is active and path is ITM
        exercised    = (exercise_val > continuation) & (exercise_val > 0)
        eta[exercised, m_idx] = 0.0

    # ── (c)  Cumulative non-exercise indicator η̃_{pn} ──────────────────────
    # η̃_{pn} = product of all η_{pm} for T_m ≤ T_n
    # Once exercised, stays zero (cumprod handles this naturally)
    eta_tilde_ex = np.cumprod(eta, axis=1)   # [M, num_ex]

    # ── (d)  Undiscounted exposure V(T_n) = Ṽ(T_n) * B(T_n) ──────────────
    # Compute numeraire B(t_i) for all time steps
    B_all = np.zeros((M, N + 1))
    for i in range(N + 1):
        r_path_i = X_paths[:, :i + 1] + cfg.r0
        integral = np.trapezoid(r_path_i, dx=dt, axis=1)
        B_all[:, i] = np.exp(integral)

    # V_undiscounted: for each exercise date, zero out paths that were exercised
    # (cash-settled: exposure vanishes after exercise)
    V_undi = Y_jumped * B_all   # [M, N+1] — undiscounted value along all paths

    # Build a time-step-level non-exercise mask [M, N+1]
    # At each time t_i, find which exercise dates have T_m ≤ t_i
    alive_mask = np.ones((M, N + 1))  # 1 = option still alive
    for m_idx, step_m in enumerate(ex_step_idx):
        # For all steps after T_m: set alive = alive * eta[:, m_idx]
        eta_m = eta[:, m_idx]                     # [M]
        # Broadcast: zero out paths that exercised at T_m for all t > T_m
        alive_mask[:, step_m + 1:] *= eta_m[:, None]

    # Apply mask: set exposure to 0 on paths that have exercised
    V_exposure = V_undi * alive_mask              # [M, N+1]

    # ── (e)  EPE and ENE ──────────────────────────────────────────────────────
    # Discounted EPE/ENE (paper eqs 2.5, 2.6)
    # EPE(T_n) = E[D(0,T_n) * max(V(T_n), 0)]
    #           = E[max(Ṽ(T_n) * alive_mask, 0)]   (already discounted via Ṽ)
    # We compute both undiscounted and discounted versions

    EPE = np.mean(np.maximum(V_exposure, 0.0), axis=0)   # [N+1]
    ENE = np.mean(np.minimum(V_exposure, 0.0), axis=0)   # [N+1]

    # ── (f)  CVA with flat hazard rate ────────────────────────────────────────
    # CVA = (1-R) * Σ_n EPE(T_n) * ΔPD(T_n)
    # ΔPD(T_n) = e^{-λ T_{n-1}} - e^{-λ T_n}  (discrete default probability)
    lam   = cfg.hazard_rate
    R     = cfg.recovery
    t_arr = time_grid[1:]                            # T_1, ..., T_N
    t_arr_prev = time_grid[:-1]                      # T_0, ..., T_{N-1}
    delta_pd   = np.exp(-lam * t_arr_prev) - np.exp(-lam * t_arr)
    CVA        = (1 - R) * np.sum(EPE[1:] * delta_pd)

    logger.info(f"  CVA = {CVA:.4f}  (λ={lam}, R={R})")

    # ── (g)  Exercise time artifacts ──────────────────────────────────────────
    # τ_p = first T_m where η_{pm} = 0, or NaN if never exercised
    exercise_times = np.full(M, np.nan)
    for m_idx, T_m in enumerate(cfg.exercise_dates):
        exercised_now = (eta[:, m_idx] == 0.0)
        # Only record first exercise (not-yet-recorded paths)
        not_yet = np.isnan(exercise_times)
        exercise_times[not_yet & exercised_now] = T_m

    # Fraction exercised at each T_m (conditional on being alive)
    exercise_prob = np.mean(eta == 0.0, axis=0)   # [num_ex]

    # Survival curve: P(not yet exercised at T_m)
    survival_curve = np.ones(num_ex)
    for m_idx in range(num_ex):
        survival_curve[m_idx] = np.mean(
            np.all(eta[:, :m_idx + 1] == 1.0, axis=1)
        )

    frac_exercised = np.mean(~np.isnan(exercise_times))
    logger.info(f"  Fraction of paths exercised: {frac_exercised:.1%}")
    logger.info(f"  Exercise probability by date:")
    for T_m, p in zip(cfg.exercise_dates, exercise_prob):
        logger.info(f"    T={T_m:.1f}y: {p:.1%}")
    logger.info("Stage 3 complete.\n")

    return {
        'time_grid':       time_grid,
        'EPE':             EPE,
        'ENE':             ENE,
        'V_undiscounted':  V_exposure,
        'exercise_times':  exercise_times,
        'exercise_prob':   exercise_prob,
        'survival_curve':  survival_curve,
        'CVA':             CVA,
        'ex_step_idx':     ex_step_idx,
        'eta':             eta,
        'Y_jumped':        Y_jumped,
        'X_paths':         X_paths,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 7.  Bachelier Fit — Reproduce Figure 3.4
# ─────────────────────────────────────────────────────────────────────────────

def bachelier_formula(x: np.ndarray, A: float, c: float, s: float) -> np.ndarray:
    """
    Bachelier-type option price formula (paper eq 3.8):
        V_Bach(x) = A * [(x-c) * Φ((x-c)/s) + s * φ((x-c)/s)]

    As x → ∞: V_Bach → A*(x-c)  (exercise value asymptote)
    Parameters:
        A : slope of exercise value
        c : intersection of exercise value with x-axis
        s : effective volatility (controls rounding near ATM)
    """
    z   = (x - c) / (s + 1e-12)
    cdf = scipy_norm.cdf(z)
    pdf = scipy_norm.pdf(z)
    return A * ((x - c) * cdf + s * pdf)


def fit_bachelier(
    x_vals: np.ndarray,
    V_vals: np.ndarray,
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Fit the Bachelier formula to (x, V) scatter at a given exercise date.

    Args:
        x_vals : x(T_m) values, shape [M]
        V_vals : Ṽ(T_m) values, shape [M]

    Returns:
        popt   : [A, c, s] optimal parameters
        V_fit  : fitted values at x_vals
    """
    # Initial guess: A from slope of exercise value (linear for HW),
    # c near zero (ATM x), s from data spread
    try:
        popt, _ = curve_fit(
            bachelier_formula,
            x_vals,
            V_vals,
            p0=[np.std(V_vals) / np.std(x_vals), 0.0, np.std(x_vals)],
            maxfev=10_000,
            bounds=([0, -np.inf, 1e-6], [np.inf, np.inf, np.inf]),
        )
    except RuntimeError:
        popt = np.array([1.0, 0.0, 0.01])   # fallback

    V_fit = bachelier_formula(x_vals, *popt)
    return popt, V_fit


# ─────────────────────────────────────────────────────────────────────────────
# 8.  Plotting — Reproduce Figures 3.1, 3.2, 3.3, 3.4
# ─────────────────────────────────────────────────────────────────────────────

def plot_epe_ene(exposure_data: Dict, cfg: Config):
    """
    Figure 3.1 (left): EPE and ENE of cash-settled Bermudan swaption.
    ENE should be ≈ 0 (receiver swaption value is always non-negative).
    EPE decreases and jumps down at exercise dates.
    """
    fig, ax = plt.subplots(figsize=(9, 5))

    t = exposure_data['time_grid']
    ax.plot(t, exposure_data['EPE'], 'b-o', markersize=2,
            linewidth=1.5, label='EPE')
    ax.plot(t, exposure_data['ENE'], 'r-o', markersize=2,
            linewidth=1.5, label='ENE')

    # Mark exercise dates
    for T_m in cfg.exercise_dates:
        ax.axvline(T_m, color='grey', linestyle=':', alpha=0.5, linewidth=0.8)

    ax.set_xlabel('Time (years)', fontsize=12)
    ax.set_ylabel('Exposure', fontsize=12)
    ax.set_title('EPE and ENE — Cash-Settled Bermudan Swaption\n'
                 '(She & Grecu 2018, Fig 3.1 left)', fontsize=12)
    ax.legend(fontsize=11)
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    return fig


def plot_loss_curve(all_losses: List[float], cfg: Config):
    """
    Figure 3.1 (right): Evolution of loss function with training steps.
    Paper shows convergence around step 500.
    """
    fig, ax = plt.subplots(figsize=(7, 5))
    steps = np.arange(1, len(all_losses) + 1)
    ax.semilogy(steps, all_losses, 'b-', linewidth=1.0, alpha=0.8)
    ax.set_xlabel('Training Step', fontsize=12)
    ax.set_ylabel('Loss (log scale)', fontsize=12)
    ax.set_title('Loss Function Evolution\n(She & Grecu 2018, Fig 3.1 right)', fontsize=12)
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    return fig


def plot_future_value_evolution(
    exposure_data: Dict,
    pre_train_data: Dict,
    Y_snapshots: List[np.ndarray],
    exercise_date_idx: int,
    cfg: Config,
):
    """
    Figures 3.2 and 3.3: Scatter of (x, V) at a given exercise date
    at training snapshots (steps 1, 100, 200, 500, 1000, 3000).

    Blue dots : portfolio value V(T_m) vs x(T_m)  ← learned
    Orange    : exercise value U(T_m) vs x(T_m)   ← fixed, analytical

    Args:
        exercise_date_idx : 0 = 1st exercise date (Fig 3.2), 3 = 4th (Fig 3.3)
    """
    step_m  = pre_train_data['ex_step_idx'][exercise_date_idx]
    T_m     = cfg.exercise_dates[exercise_date_idx]
    x_Tm    = pre_train_data['X_paths'][:, step_m]
    U_Tm    = pre_train_data['U_exercise'][:, exercise_date_idx]

    # Use a subset of snapshots for the panels
    snap_labels = [f'iter {i+1}' for i in range(len(Y_snapshots))]
    n_snaps     = len(Y_snapshots)

    n_cols = min(3, n_snaps)
    n_rows = math.ceil(n_snaps / n_cols)
    fig, axes = plt.subplots(n_rows, n_cols,
                              figsize=(5 * n_cols, 4 * n_rows))
    axes = np.array(axes).flatten()

    for k, (Y_snap, label) in enumerate(zip(Y_snapshots, snap_labels)):
        ax = axes[k]
        # Convert discounted Ṽ back to undiscounted V for plotting
        B_Tm   = pre_train_data['B_exercise'][:, exercise_date_idx]
        V_Tm   = Y_snap[:, step_m] * B_Tm

        ax.scatter(x_Tm, V_Tm, s=2, alpha=0.3, color='steelblue', label='V (learned)')
        ax.scatter(x_Tm, U_Tm, s=2, alpha=0.3, color='orange',    label='U (exercise)')
        ax.set_title(label, fontsize=10)
        ax.set_xlabel('x', fontsize=9)
        ax.set_ylabel('V', fontsize=9)
        ax.grid(True, alpha=0.2)
        if k == 0:
            ax.legend(fontsize=8)

    # Hide empty panels
    for k in range(n_snaps, len(axes)):
        axes[k].set_visible(False)

    fig.suptitle(
        f'Future Value Evolution — Exercise Date {exercise_date_idx+1} '
        f'(T={T_m:.1f}y)\n(She & Grecu 2018, Fig 3.{"2" if exercise_date_idx==0 else "3"})',
        fontsize=12,
    )
    plt.tight_layout()
    return fig


def plot_bachelier_fit(
    exposure_data: Dict,
    pre_train_data: Dict,
    exercise_date_idx: int,
    cfg: Config,
):
    """
    Figure 3.4: Fit of V(x) at an exercise date to the Bachelier formula.
    Blue dots: (x, V) scatter. Red dashed: Bachelier fit.
    """
    step_m = pre_train_data['ex_step_idx'][exercise_date_idx]
    T_m    = cfg.exercise_dates[exercise_date_idx]

    x_Tm   = pre_train_data['X_paths'][:, step_m]
    Y_snap = exposure_data['Y_jumped']
    B_Tm   = pre_train_data['B_exercise'][:, exercise_date_idx]
    V_Tm   = Y_snap[:, step_m] * B_Tm   # undiscounted

    # Fit Bachelier on the ITM / near-ATM region
    popt, V_fit = fit_bachelier(x_Tm, V_Tm)
    A, c, s = popt

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.scatter(x_Tm, V_Tm, s=3, alpha=0.3, color='blue', label='V (learned)')

    # Plot fitted curve on sorted x grid
    x_sorted = np.sort(x_Tm)
    V_sorted = bachelier_formula(x_sorted, A, c, s)
    ax.plot(x_sorted, V_sorted, 'r--', linewidth=2,
            label=f'Bachelier fit\nA={A:.3f}, c={c:.4f}, s={s:.4f}')

    ax.set_xlabel('x', fontsize=12)
    ax.set_ylabel('V', fontsize=12)
    ax.set_title(
        f'Bachelier Fit — Exercise Date {exercise_date_idx+1} (T={T_m:.1f}y)\n'
        f'(She & Grecu 2018, Fig 3.4)',
        fontsize=12,
    )
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    return fig


def plot_exercise_time_distribution(exposure_data: Dict, cfg: Config):
    """
    Exercise time artifacts:
        (a) Histogram of exercise times τ_p across paths
        (b) Exercise probability at each T_m
        (c) Survival curve P(not yet exercised)
    """
    exercise_times = exposure_data['exercise_times']
    exercise_prob  = exposure_data['exercise_prob']
    survival_curve = exposure_data['survival_curve']
    ex_dates       = cfg.exercise_dates

    fig, axes = plt.subplots(1, 3, figsize=(15, 5))

    # (a) Histogram
    exercised = exercise_times[~np.isnan(exercise_times)]
    axes[0].hist(exercised, bins=len(ex_dates),
                 color='steelblue', edgecolor='black', alpha=0.7)
    axes[0].set_xlabel('Exercise Time (years)', fontsize=11)
    axes[0].set_ylabel('Number of Paths', fontsize=11)
    axes[0].set_title('Exercise Time Distribution', fontsize=12)
    axes[0].grid(True, alpha=0.3)
    pct = 100 * len(exercised) / len(exercise_times)
    axes[0].text(0.05, 0.95, f'{pct:.1f}% exercised',
                 transform=axes[0].transAxes, fontsize=10,
                 verticalalignment='top')

    # (b) Exercise probability at each T_m
    axes[1].bar(ex_dates, exercise_prob, width=0.3,
                color='orange', edgecolor='black', alpha=0.8)
    axes[1].set_xlabel('Exercise Date (years)', fontsize=11)
    axes[1].set_ylabel('Probability', fontsize=11)
    axes[1].set_title('Exercise Probability by Date', fontsize=12)
    axes[1].set_ylim(0, 1)
    axes[1].grid(True, alpha=0.3)

    # (c) Survival curve
    axes[2].step(ex_dates, survival_curve, where='post',
                 color='green', linewidth=2)
    axes[2].scatter(ex_dates, survival_curve, color='green', s=60, zorder=3)
    axes[2].set_xlabel('Time (years)', fontsize=11)
    axes[2].set_ylabel('P(not yet exercised)', fontsize=11)
    axes[2].set_title('Survival Curve', fontsize=12)
    axes[2].set_ylim(0, 1.05)
    axes[2].grid(True, alpha=0.3)

    plt.suptitle('Exercise Time Artifacts', fontsize=13, fontweight='bold')
    plt.tight_layout()
    return fig


# ─────────────────────────────────────────────────────────────────────────────
# 9.  Excel Export
# ─────────────────────────────────────────────────────────────────────────────

def export_to_excel(
    cfg: Config,
    pre_train_data: Dict,
    exposure_data: Dict,
    all_losses: List[float],
    output_path: str = "bermudan_swaption_results.xlsx",
):
    """
    Export all results to a formatted Excel workbook with four sheets:

    Sheet 1 — Exposure Profiles  : EPE, ENE at every time step
    Sheet 2 — Exercise Times      : per-path exercise times + summary
    Sheet 3 — Exercise Analytics  : probability by date, survival curve
    Sheet 4 — Training Loss       : loss at every training step
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Build DataFrames ──────────────────────────────────────────────────────

    # Sheet 1: Exposure profiles (sampled at exercise dates + a few extras)
    t_grid = exposure_data['time_grid']
    EPE    = exposure_data['EPE']
    ENE    = exposure_data['ENE']

    df_exposure = pd.DataFrame({
        'Time (years)': t_grid,
        'EPE':          EPE,
        'ENE':          ENE,
        'Discount Factor': np.exp(-cfg.r0 * t_grid),
    })

    # Sheet 2: Per-path exercise times
    ex_times = exposure_data['exercise_times']
    df_exercise = pd.DataFrame({
        'Path':          np.arange(cfg.num_paths),
        'Exercise Time': ex_times,
        'Exercised':     (~np.isnan(ex_times)).astype(int),
    })

    # Sheet 3: Exercise analytics
    ex_dates       = cfg.exercise_dates
    exercise_prob  = exposure_data['exercise_prob']
    survival_curve = exposure_data['survival_curve']
    df_analytics = pd.DataFrame({
        'Exercise Date (years)': ex_dates,
        'Exercise Probability':  exercise_prob,
        'Survival Probability':  survival_curve,
    })

    # Sheet 4: Training loss
    df_loss = pd.DataFrame({
        'Training Step': np.arange(1, len(all_losses) + 1),
        'MSE Loss':      all_losses,
    })

    # ── Write to Excel ────────────────────────────────────────────────────────
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_exposure.to_excel(writer,  sheet_name='Exposure Profiles',  index=False)
        df_exercise.to_excel(writer,  sheet_name='Exercise Times',      index=False)
        df_analytics.to_excel(writer, sheet_name='Exercise Analytics',  index=False)
        df_loss.to_excel(writer,      sheet_name='Training Loss',        index=False)

        # ── Formatting ────────────────────────────────────────────────────────
        wb = writer.book

        header_font    = Font(bold=True, color='FFFFFF')
        header_fill    = PatternFill('solid', fgColor='1F4E79')  # dark blue
        num_fill       = PatternFill('solid', fgColor='EBF3FB')  # light blue
        center_align   = Alignment(horizontal='center')
        thin_border    = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin'),
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Header row formatting
            for cell in ws[1]:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align
                cell.border    = thin_border

            # Data rows: alternating fill + borders
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                fill = num_fill if row_idx % 2 == 0 else PatternFill()
                for cell in row:
                    cell.fill   = fill
                    cell.border = thin_border
                    if isinstance(cell.value, float):
                        cell.number_format = '0.0000'

            # Auto-width columns
            for col in ws.columns:
                max_len = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col
                )
                ws.column_dimensions[get_column_letter(col[0].column)].width = (
                    min(max_len + 4, 25)
                )

        # ── Summary sheet ─────────────────────────────────────────────────────
        ws_sum = wb.create_sheet('Summary', 0)
        rows = [
            ('Parameter', 'Value'),
            ('Notional',             cfg.notional),
            ('Fixed Rate',           f'{cfg.fixed_rate:.2%}'),
            ('Swap Tenor (years)',    cfg.swap_tenor),
            ('Exercise Start',        cfg.exercise_start),
            ('Exercise End',          cfg.exercise_end),
            ('kappa',                 cfg.kappa),
            ('sigma_r',               cfg.sigma_r),
            ('r0',                    f'{cfg.r0:.2%}'),
            ('Num Paths',             cfg.num_paths),
            ('Num Time Steps',        cfg.N),
            ('dt',                    f'{cfg.dt:.5f}'),
            ('Hazard Rate',           f'{cfg.hazard_rate:.2%}'),
            ('Recovery Rate',         f'{cfg.recovery:.0%}'),
            ('CVA',                   f'{exposure_data["CVA"]:.4f}'),
            ('% Paths Exercised',
             f'{100*np.mean(~np.isnan(exposure_data["exercise_times"])):.1f}%'),
            ('Final Training Loss',   f'{all_losses[-1]:.4e}'),
        ]
        for row in rows:
            ws_sum.append(row)

        # Header
        for cell in ws_sum[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
        for col in ws_sum.columns:
            ws_sum.column_dimensions[get_column_letter(col[0].column)].width = 28

    logger.info(f"Results exported to: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# 10.  Main Orchestrator
# ─────────────────────────────────────────────────────────────────────────────

def main():
    """
    Full pipeline:
        1. Configure
        2. Pre-training  (simulate paths + exercise values)
        3. Training      (Picard iteration with exercise jumps)
        4. Exposure      (EPE, ENE, CVA, exercise times)
        5. Plots         (Figs 3.1–3.4)
        6. Export        (Excel)
    """
    t_start = time.time()

    # ── 1.  Configuration ─────────────────────────────────────────────────────
    cfg = Config(
        num_paths      = 5_000,
        num_train_steps= 2_000,   # 5000 for full paper convergence; 2000 for speed
        log_every      = 500,
    )
    hw = HullWhite(cfg)

    logger.info("Bermudan Swaption CVA — She & Grecu (2018) Section 3")
    logger.info(f"Exercise dates: {cfg.exercise_dates}")
    logger.info(f"T={cfg.T:.1f}y, N={cfg.N} steps, dt={cfg.dt:.4f}")
    logger.info(f"Paths={cfg.num_paths}, Train steps={cfg.num_train_steps}\n")

    # ── 2.  Stage 1: Pre-training ─────────────────────────────────────────────
    pre_data = pre_training(cfg, hw)

    # ── 3.  Stage 2: Training ─────────────────────────────────────────────────
    net, all_losses, Y_snapshots = train(
        cfg, hw, pre_data, num_picard_iters=3
    )

    # ── 4.  Stage 3: Exposure ─────────────────────────────────────────────────
    exp_data = compute_exposure(cfg, hw, net, pre_data)

    logger.info(f"CVA = {exp_data['CVA']:.6f}")
    logger.info(f"EPE at t=0: {exp_data['EPE'][0]:.4f}")

    # ── 5.  Plots ─────────────────────────────────────────────────────────────

    # Fig 3.1 left: EPE / ENE
    fig_epe = plot_epe_ene(exp_data, cfg)
    fig_epe.savefig('fig3_1_epe_ene.png', dpi=150, bbox_inches='tight')

    # Fig 3.1 right: Loss curve
    fig_loss = plot_loss_curve(all_losses, cfg)
    fig_loss.savefig('fig3_1_loss.png', dpi=150, bbox_inches='tight')

    # Fig 3.2: Future value evolution at 1st exercise date
    fig_fv1 = plot_future_value_evolution(
        exp_data, pre_data, Y_snapshots, exercise_date_idx=0, cfg=cfg
    )
    fig_fv1.savefig('fig3_2_fv_ex1.png', dpi=150, bbox_inches='tight')

    # Fig 3.3: Future value evolution at 4th exercise date (if exists)
    if len(cfg.exercise_dates) >= 4:
        fig_fv4 = plot_future_value_evolution(
            exp_data, pre_data, Y_snapshots, exercise_date_idx=3, cfg=cfg
        )
        fig_fv4.savefig('fig3_3_fv_ex4.png', dpi=150, bbox_inches='tight')

    # Fig 3.4: Bachelier fit at 1st exercise date
    fig_bach = plot_bachelier_fit(exp_data, pre_data, exercise_date_idx=0, cfg=cfg)
    fig_bach.savefig('fig3_4_bachelier.png', dpi=150, bbox_inches='tight')

    # Exercise time artifacts
    fig_ex = plot_exercise_time_distribution(exp_data, cfg)
    fig_ex.savefig('exercise_time_distribution.png', dpi=150, bbox_inches='tight')

    # ── 6.  Export to Excel ───────────────────────────────────────────────────
    export_to_excel(cfg, pre_data, exp_data, all_losses,
                    output_path='bermudan_swaption_results.xlsx')

    elapsed = time.time() - t_start
    logger.info(f"\nTotal runtime: {elapsed:.1f}s")
    logger.info("Done.")

    return cfg, pre_data, net, all_losses, Y_snapshots, exp_data


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    cfg, pre_data, net, all_losses, Y_snapshots, exp_data = main()
